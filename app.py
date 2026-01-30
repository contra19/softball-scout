"""
Softball Scout Stats - Streamlit App (v2)

A comprehensive softball statistics management app with:
- Season tracking (batting, pitching, game results)
- Team scouting
- Player analytics and comparisons
- Excel import/export
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
import tempfile
import os
import hashlib
import extra_streamlit_components as stx
import database as db
from importer import import_file, parse_filename_for_game_info, preview_csv


# =============================================================================
# PAGE CONFIG
# =============================================================================

st.set_page_config(
    page_title="Softball Scout Stats",
    page_icon="🥎",
    layout="wide",
    initial_sidebar_state="expanded"
)


# =============================================================================
# AUTHENTICATION
# =============================================================================

def get_cookie_manager():
    """Get or create the cookie manager instance in session state."""
    if "cookie_manager" not in st.session_state:
        st.session_state.cookie_manager = stx.CookieManager()
    return st.session_state.cookie_manager

def generate_auth_token(password: str) -> str:
    """Generate a hash token for the password to store in cookie."""
    return hashlib.sha256(password.encode()).hexdigest()

def check_password():
    """Returns True if the user has entered the correct password.
    Uses cookies to persist authentication across page refreshes.
    """
    cookie_manager = get_cookie_manager()
    correct_password = st.secrets.get("password", "softball")
    expected_token = generate_auth_token(correct_password)

    def password_entered():
        """Checks whether the password entered is correct."""
        st.session_state["login_attempted"] = True
        entered_password = st.session_state.get("password", "")
        if entered_password == correct_password:
            st.session_state["authenticated"] = True
            # Set cookie for 30 days
            cookie_manager.set("auth_token", expected_token, expires_at=datetime.now() + timedelta(days=30))
            del st.session_state["password"]  # Don't store password
        else:
            st.session_state["authenticated"] = False

    # First run - initialize state
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False
        st.session_state["login_attempted"] = False
        st.session_state["cookie_checked"] = False

    # Check for existing auth cookie
    # The cookie manager needs a rerun to read cookies on first load
    if not st.session_state["authenticated"]:
        auth_cookie = cookie_manager.get("auth_token")
        if auth_cookie == expected_token:
            st.session_state["authenticated"] = True
        elif auth_cookie is None and not st.session_state.get("cookie_checked"):
            # Cookie manager returns None on first render before JS loads
            # Mark that we've attempted to check and trigger a rerun
            st.session_state["cookie_checked"] = True
            st.rerun()

    if not st.session_state["authenticated"]:
        st.title("🥎 Softball Scout Stats")
        st.write("Please enter the password to access the app.")
        st.text_input(
            "Password",
            type="password",
            key="password",
            on_change=password_entered
        )
        # Only show error if a login was actually attempted and failed
        if st.session_state.get("login_attempted") and not st.session_state["authenticated"]:
            st.error("Incorrect password")
        return False

    return True


# =============================================================================
# DIALOG DEFINITIONS
# =============================================================================

@st.dialog("Add Game")
def add_game_dialog():
    """Dialog for adding a new game"""
    if not st.session_state.get('current_season_id'):
        st.warning("No season selected")
        return

    season_id = st.session_state.current_season_id

    # Generate time options (6 AM to 10 PM in 5-minute intervals)
    time_options = [""]
    for hour in range(6, 22):  # 6 AM to 9 PM
        for minute in range(0, 60, 5):
            h12 = hour if hour <= 12 else hour - 12
            if h12 == 0:
                h12 = 12
            ampm = "AM" if hour < 12 else "PM"
            time_options.append(f"{h12}:{minute:02d} {ampm}")

    st.subheader("Game Information")
    col1, col2 = st.columns(2)
    with col1:
        # Game Date picker
        add_date_picker = st.date_input(
            "Game Date *",
            value="today",
            key="add_date_picker"
        )
        game_date = f"{add_date_picker.month}/{add_date_picker.day}" if add_date_picker else ""

        # Game Time - single dropdown with all options
        game_time = st.selectbox("Game Time", options=time_options, key="add_game_time")

        opponent = st.text_input("Opponent *", placeholder="e.g., NJ Vipers 12U")
    with col2:
        win_loss = st.selectbox("Result", ["", "W", "L", "T"])
        rf_col, ra_col = st.columns(2)
        with rf_col:
            runs_for = st.number_input("Runs For", min_value=0, value=0)
        with ra_col:
            runs_against = st.number_input("Runs Against", min_value=0, value=0)

    if st.button("Save Game", type="primary"):
        if not game_date or not opponent:
            st.error("Game date and opponent are required")
            return

        db.create_game(
            season_id=season_id,
            game_date=game_date,
            game_time=game_time if game_time else None,
            opponent_name=opponent,
            win_loss=win_loss if win_loss else None,
            runs_for=runs_for if runs_for > 0 else None,
            runs_against=runs_against if runs_against > 0 else None
        )
        st.success(f"Game added: {game_date} {game_time or ''} vs {opponent}".strip())
        st.rerun()


@st.dialog("Edit Game")
def edit_game_dialog(game_id: int):
    """Dialog for editing an existing game"""
    game = db.get_game(game_id)
    if not game:
        st.error("Game not found")
        return

    st.subheader("Edit Game Information")

    # Generate time options (6 AM to 10 PM in 5-minute intervals)
    time_options = [""]
    for hour in range(6, 22):  # 6 AM to 9 PM
        for minute in range(0, 60, 5):
            h12 = hour if hour <= 12 else hour - 12
            if h12 == 0:
                h12 = 12
            ampm = "AM" if hour < 12 else "PM"
            time_options.append(f"{h12}:{minute:02d} {ampm}")

    # Find index of existing time
    time_index = 0
    if game.game_time and game.game_time in time_options:
        time_index = time_options.index(game.game_time)

    col1, col2 = st.columns(2)
    with col1:
        # Game Date picker
        edit_date_picker = st.date_input(
            "Game Date *",
            value="today",
            key="edit_date_picker"
        )
        game_date = f"{edit_date_picker.month}/{edit_date_picker.day}" if edit_date_picker else game.game_date

        # Game Time - single dropdown with all options
        game_time = st.selectbox("Game Time", options=time_options, index=time_index, key="edit_game_time")

        opponent = st.text_input("Opponent *", value=game.opponent_name or '', key="edit_opponent")
    with col2:
        wl_options = ["", "W", "L", "T"]
        wl_index = wl_options.index(game.win_loss) if game.win_loss in wl_options else 0
        win_loss = st.selectbox("Result", wl_options, index=wl_index, key="edit_wl")
        runs_for = st.number_input("Runs For", min_value=0, value=game.runs_for or 0, key="edit_rf")
        runs_against = st.number_input("Runs Against", min_value=0, value=game.runs_against or 0, key="edit_ra")

    if st.button("Save Changes", type="primary", key="save_game_changes"):
        if not game_date or not opponent:
            st.error("Game date and opponent are required")
            return

        db.update_game(
            game_id=game_id,
            game_date=game_date,
            game_time=game_time if game_time else None,
            opponent_name=opponent,
            win_loss=win_loss if win_loss else None,
            runs_for=runs_for if runs_for > 0 else None,
            runs_against=runs_against if runs_against > 0 else None
        )
        st.success("Game updated!")
        st.rerun()


@st.dialog("Edit Batting Stats", width="large")
def edit_batting_stats_dialog(game_id: int):
    """Dialog for editing batting stats for a game"""
    game = db.get_game(game_id)
    if not game:
        st.error("Game not found")
        return

    st.subheader(f"Edit Batting Stats: {game.game_date} vs {game.opponent_name}")

    batting = db.get_batting_stats_for_game(game_id)
    if not batting:
        st.info("No batting stats recorded for this game.")
        return

    st.write("Edit the stats below and click Save to update:")

    # Create editable dataframe
    edited_stats = []
    for stat in batting:
        st.write(f"**{stat['player_name']}**")
        cols = st.columns(8)
        with cols[0]:
            ab = st.number_input("AB", value=stat.get('ab', 0), min_value=0, key=f"ab_{stat['id']}")
        with cols[1]:
            r = st.number_input("R", value=stat.get('r', 0), min_value=0, key=f"r_{stat['id']}")
        with cols[2]:
            h = st.number_input("H", value=stat.get('h', 0), min_value=0, key=f"h_{stat['id']}")
        with cols[3]:
            rbi = st.number_input("RBI", value=stat.get('rbi', 0), min_value=0, key=f"rbi_{stat['id']}")
        with cols[4]:
            bb = st.number_input("BB", value=stat.get('bb', 0), min_value=0, key=f"bb_{stat['id']}")
        with cols[5]:
            so = st.number_input("SO", value=stat.get('so', 0), min_value=0, key=f"so_{stat['id']}")
        with cols[6]:
            hbp = st.number_input("HBP", value=stat.get('hbp', 0), min_value=0, key=f"hbp_{stat['id']}")
        with cols[7]:
            sac = st.number_input("SAC", value=stat.get('sac', 0), min_value=0, key=f"sac_{stat['id']}")

        edited_stats.append({
            'id': stat['id'],
            'ab': ab, 'r': r, 'h': h, 'rbi': rbi,
            'bb': bb, 'so': so, 'hbp': hbp, 'sac': sac
        })
        st.divider()

    if st.button("Save All Changes", type="primary", key="save_batting_changes"):
        for stat in edited_stats:
            db.update_batting_stats(
                batting_stats_id=stat['id'],
                ab=stat['ab'], r=stat['r'], h=stat['h'], rbi=stat['rbi'],
                bb=stat['bb'], so=stat['so'], hbp=stat['hbp'], sac=stat['sac']
            )
        st.success("Batting stats updated!")
        st.rerun()


@st.dialog("Add Player")
def add_player_dialog():
    """Dialog for adding a new player"""
    st.subheader("Player Information")

    col1, col2 = st.columns(2)
    with col1:
        name = st.text_input("Name *")
        jersey = st.text_input("Jersey #")
    with col2:
        bats = st.selectbox("Bats", ["", "R", "L", "S"])
        throws = st.selectbox("Throws", ["", "R", "L"])

    if st.button("Save Player", type="primary"):
        if not name:
            st.error("Name is required")
            return

        db.create_player(name, jersey if jersey else None,
                        bats if bats else None, throws if throws else None)
        st.success(f"Added player: {name}")
        st.rerun()


@st.dialog("Add Opponent Team")
def add_opponent_team_dialog():
    """Dialog for adding an opponent team for scouting"""
    st.subheader("Team Information")

    name = st.text_input("Team Name *")
    location = st.text_input("Location")

    if st.button("Save Team", type="primary"):
        if not name:
            st.error("Team name is required")
            return

        db.create_team(name, location if location else None)
        st.success(f"Added team: {name}")
        st.rerun()


@st.dialog("Add Age Group")
def add_age_group_dialog():
    """Dialog for adding an age group"""
    st.subheader("Age Group")

    name = st.text_input("Name *", placeholder="e.g., 12U, 10U, 14U")

    if st.button("Save Age Group", type="primary"):
        if not name:
            st.error("Name is required")
            return

        db.get_or_create_age_group(name)
        st.success(f"Added age group: {name}")
        st.rerun()


@st.dialog("Add Our Team")
def add_our_team_dialog():
    """Dialog for adding one of our teams"""
    st.subheader("Team Information")

    age_groups = db.get_all_age_groups()
    if not age_groups:
        st.warning("Create an age group first")
        return

    col1, col2 = st.columns(2)
    with col1:
        name = st.text_input("Team Name *", placeholder="e.g., PA Chaos Taranto")
        location = st.text_input("Location", placeholder="e.g., Pennsylvania")
    with col2:
        ag_options = {ag.name: ag.id for ag in age_groups}
        selected_ag = st.selectbox("Age Group *", options=list(ag_options.keys()))

    if st.button("Save Team", type="primary"):
        if not name or not selected_ag:
            st.error("Name and age group are required")
            return

        age_group_id = ag_options[selected_ag]
        db.get_or_create_our_team(name, age_group_id, location if location else None)
        st.success(f"Added team: {name}")
        st.rerun()


@st.dialog("Import Excel", width="large")
def import_excel_dialog():
    """Dialog for importing Excel data"""
    st.write("""
    Import data from an existing Excel workbook in the coach's format.
    This will import seasons, games, and player stats.
    """)

    uploaded_file = st.file_uploader("Upload Excel file", type=['xlsx'])

    if uploaded_file and st.button("Import Data", type="primary"):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name

        try:
            with st.spinner("Importing..."):
                results = import_file(tmp_path, verbose=False)

            st.success("Import completed!")
            st.write(f"- Sheets processed: {results['sheets_processed']}")
            st.write(f"- Games imported: {results['total_games']}")
            st.write(f"- Batting records: {results['total_batting']}")
            st.write(f"- Pitching records: {results['total_pitching']}")

            if st.button("Close"):
                st.rerun()

        except Exception as e:
            st.error(f"Import failed: {e}")
            import traceback
            st.code(traceback.format_exc())

        finally:
            os.unlink(tmp_path)


@st.dialog("Import GameChanger CSV", width="large")
def import_gamechanger_dialog():
    """Dialog for importing GameChanger CSV files"""
    if not st.session_state.get('current_season_id'):
        st.warning("Please select a season first from the sidebar.")
        return

    season_id = st.session_state.current_season_id
    season = db.get_season(season_id)
    if not season:
        st.error("Selected season not found")
        return

    st.info(f"Importing to: **{season.season_type} {season.year}**")

    st.write("""
    Upload GameChanger CSV exports. Name files like `game1_sep6_vipers.csv`
    to auto-detect date and opponent.
    """)

    uploaded_files = st.file_uploader(
        "Upload CSV file(s)",
        type=['csv'],
        accept_multiple_files=True,
        key="gc_csv_dialog"
    )

    if uploaded_files:
        file_info = []
        for f in uploaded_files:
            info = parse_filename_for_game_info(f.name)
            file_info.append({
                'file': f,
                'name': f.name,
                'game_date': info.get('game_date') or 'Unknown',
                'opponent': info.get('opponent') or f.name.replace('.csv', '')
            })

        preview_df = pd.DataFrame([{
            'File': fi['name'],
            'Date': fi['game_date'],
            'Opponent': fi['opponent']
        } for fi in file_info])
        st.dataframe(preview_df, width='stretch', hide_index=True)

        col1, col2 = st.columns(2)
        with col1:
            override_date = st.text_input("Override Date", placeholder="Leave blank for auto")
        with col2:
            override_opponent = st.text_input("Override Opponent", placeholder="Leave blank for auto")

        if st.button("Import Games", type="primary"):
            total_stats = 0
            games_created = 0

            for fi in file_info:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
                    tmp.write(fi['file'].getvalue())
                    tmp_path = tmp.name

                try:
                    game_date = override_date if override_date else fi['game_date']
                    opponent = override_opponent if override_opponent else fi['opponent']

                    result = import_file(
                        tmp_path, season_id=season_id,
                        game_date=game_date, opponent=opponent, verbose=False
                    )

                    if result.get('game_id'):
                        games_created += 1
                        total_stats += result.get('stats_imported', 0)
                        st.success(f"Imported: {game_date} vs {opponent}")

                except Exception as e:
                    st.error(f"Error: {fi['name']}: {e}")

                finally:
                    os.unlink(tmp_path)

            st.divider()
            st.success(f"Created {games_created} game(s) with {total_stats} player stats.")
            if st.button("Close"):
                st.rerun()


# =============================================================================
# SIDEBAR NAVIGATION
# =============================================================================

def render_sidebar():
    """Render the sidebar navigation with hierarchical selectors"""
    st.sidebar.title("🥎 Softball Scout")

    # Age Group selector
    age_groups = db.get_all_age_groups()
    if age_groups:
        age_group_options = {ag.name: ag.id for ag in age_groups}
        selected_age_group = st.sidebar.selectbox(
            "Age Group",
            options=list(age_group_options.keys()),
            key="selected_age_group"
        )
        st.session_state.current_age_group_id = age_group_options.get(selected_age_group)
    else:
        st.sidebar.info("No age groups yet. Go to Setup.")
        st.session_state.current_age_group_id = None

    # Team selector (filtered by age group) - with "All Teams" option
    if st.session_state.get('current_age_group_id'):
        teams = db.get_all_our_teams(age_group_id=st.session_state.current_age_group_id)
        if teams:
            # Build options list with "All Teams" first
            team_names = ["All Teams"] + [t.name for t in teams]
            team_id_map = {t.name: t.id for t in teams}

            selected_team = st.sidebar.selectbox(
                "Team",
                options=team_names,
                key="selected_team"
            )

            if selected_team == "All Teams":
                st.session_state.current_team_id = None
            else:
                st.session_state.current_team_id = team_id_map.get(selected_team)
        else:
            st.sidebar.info("No teams in this age group.")
            st.session_state.current_team_id = None
    else:
        st.session_state.current_team_id = None

    # Season selector - with "All Seasons" option
    if st.session_state.get('current_age_group_id'):
        # Get available seasons based on team selection
        if st.session_state.get('current_team_id'):
            # Specific team selected - show that team's seasons
            team_seasons = db.get_all_seasons(team_id=st.session_state.current_team_id)
            season_names = ["All Seasons"] + [f"{s.season_type} {s.year}" for s in team_seasons]
            season_id_map = {f"{s.season_type} {s.year}": s.id for s in team_seasons}
        else:
            # "All Teams" selected - show unique season types for the age group
            age_group_seasons = db.get_available_seasons_for_age_group(
                st.session_state.current_age_group_id)
            season_names = ["All Seasons"] + [f"{s['season_type']} {s['year']}" for s in age_group_seasons]
            # For "All Teams", store (year, season_type) tuple
            season_id_map = {f"{s['season_type']} {s['year']}": (s['year'], s['season_type'])
                            for s in age_group_seasons}

        if len(season_names) > 1:  # More than just "All Seasons"
            selected_season = st.sidebar.selectbox(
                "Season",
                options=season_names,
                key="selected_season"
            )

            if selected_season == "All Seasons":
                st.session_state.current_season_id = None
                st.session_state.current_season_info = None
            else:
                value = season_id_map.get(selected_season)
                if isinstance(value, tuple):
                    # "All Teams" mode - store tuple
                    st.session_state.current_season_id = None
                    st.session_state.current_season_info = value  # (year, season_type)
                else:
                    # Specific team mode - store season id
                    st.session_state.current_season_id = value
                    st.session_state.current_season_info = None
        else:
            st.sidebar.info("No seasons yet.")
            st.session_state.current_season_id = None
            st.session_state.current_season_info = None
    else:
        st.session_state.current_season_id = None
        st.session_state.current_season_info = None

    st.sidebar.divider()

    # Navigation
    page = st.sidebar.radio(
        "Navigation",
        ["Dashboard", "Games", "Players", "Teams", "Setup"],
        key="nav_page"
    )

    return page


# =============================================================================
# DASHBOARD PAGE
# =============================================================================

def render_dashboard():
    """Main dashboard based on sidebar Team/Season selections"""
    st.title("📊 Dashboard")

    if not st.session_state.get('current_age_group_id'):
        st.warning("No age group selected. Go to Setup to import data or create an age group.")
        return

    age_group_id = st.session_state.current_age_group_id
    age_group = db.get_age_group(age_group_id)

    if not age_group:
        st.error("Age group not found")
        return

    # Get current selections from sidebar
    team_id = st.session_state.get('current_team_id')  # None = "All Teams"
    season_id = st.session_state.get('current_season_id')  # None if "All Seasons" or "All Teams" mode
    season_info = st.session_state.get('current_season_info')  # (year, type) tuple for "All Teams" mode

    # Determine what to display based on selections
    if team_id is None:
        # "All Teams" selected - show league-wide stats
        render_league_dashboard(age_group, season_info)
    else:
        # Specific team selected
        render_team_dashboard(team_id, season_id)


def render_league_dashboard(age_group, season_info):
    """League-wide dashboard showing stats across all teams in the age group.

    Args:
        age_group: The age group object
        season_info: Either (year, season_type) tuple for specific season, or None for all seasons
    """
    if season_info:
        year, season_type = season_info
        st.header(f"{age_group.name} League - {season_type} {year}")
        # Get stats for specific season
        totals = db.get_league_totals_by_season(age_group.id, year, season_type)
        batting_stats = db.get_league_batting_stats_by_season(age_group.id, year, season_type)
        pitching_stats = db.get_league_pitching_stats_by_season(age_group.id, year, season_type)
        top_batters = db.get_league_top_batters_by_season(age_group.id, year, season_type, 'ba', limit=5)
        top_pitchers = db.get_league_top_pitchers_by_season(age_group.id, year, season_type, 'era', limit=5)
    else:
        st.header(f"{age_group.name} League - All Seasons")
        # Get stats across all seasons
        totals = db.get_league_totals(age_group.id)
        batting_stats = db.get_league_batting_stats(age_group.id)
        pitching_stats = db.get_league_pitching_stats(age_group.id)
        top_batters = db.get_league_top_batters(age_group.id, 'ba', limit=5)
        top_pitchers = db.get_league_top_pitchers(age_group.id, 'era', limit=5)

    if not totals or not totals.get('games'):
        st.info("No stats available for this selection yet.")
        return

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Teams", totals.get('teams') or 0)
    with col2:
        st.metric("Games", totals.get('games') or 0)
    with col3:
        st.metric("Players", len(batting_stats) if batting_stats else 0)

    st.divider()

    # Two columns: Batting and Pitching leaders
    col_bat, col_pitch = st.columns(2)

    with col_bat:
        st.subheader("🏏 Top Batters (BA)")
        if top_batters:
            df = pd.DataFrame(top_batters)
            df = df[['player_name', 'team_name', 'ab', 'h', 'r', 'rbi', 'ba']]
            df.columns = ['Player', 'Team', 'AB', 'H', 'R', 'RBI', 'BA']
            st.dataframe(df, width='stretch', hide_index=True)
        else:
            st.info("No batting stats yet (min 10 AB)")

    with col_pitch:
        st.subheader("⚾ Top Pitchers (ERA)")
        if top_pitchers:
            df = pd.DataFrame(top_pitchers)
            df = df[['player_name', 'team_name', 'ip', 'k', 'bb', 'era', 'whip']]
            df.columns = ['Player', 'Team', 'IP', 'K', 'BB', 'ERA', 'WHIP']
            st.dataframe(df, width='stretch', hide_index=True)
        else:
            st.info("No pitching stats yet (min 5 IP)")

    st.divider()

    # Full batting stats
    st.subheader("📋 All Batting Stats")
    if batting_stats:
        df = pd.DataFrame(batting_stats)
        df = df[['player_name', 'team_name', 'jersey_number', 'ab', 'r', 'h', 'rbi', 'bb', 'so', 'ba']]
        df.columns = ['Player', 'Team', '#', 'AB', 'R', 'H', 'RBI', 'BB', 'SO', 'BA']
        st.dataframe(df, width='stretch', hide_index=True)

        total_ab = sum(s['ab'] or 0 for s in batting_stats)
        total_r = sum(s['r'] or 0 for s in batting_stats)
        total_h = sum(s['h'] or 0 for s in batting_stats)
        total_rbi = sum(s['rbi'] or 0 for s in batting_stats)
        total_bb = sum(s['bb'] or 0 for s in batting_stats)
        total_so = sum(s['so'] or 0 for s in batting_stats)
        league_ba = total_h / total_ab if total_ab > 0 else 0

        st.write(f"**League Totals:** {total_ab} AB, {total_r} R, {total_h} H, "
                 f"{total_rbi} RBI, {total_bb} BB, {total_so} SO, {league_ba:.3f} BA")

    st.divider()

    # Full pitching stats
    st.subheader("📋 All Pitching Stats")
    if pitching_stats:
        df = pd.DataFrame(pitching_stats)
        df = df[['player_name', 'team_name', 'jersey_number', 'app', 'ip', 'h', 'r', 'k', 'bb', 'era', 'whip', 'strike_pct']]
        df.columns = ['Player', 'Team', '#', 'APP', 'IP', 'H', 'R', 'K', 'BB', 'ERA', 'WHIP', 'S%']
        df['S%'] = df['S%'].apply(lambda x: f"{x:.1%}" if x else "0.0%")
        st.dataframe(df, width='stretch', hide_index=True)


def render_team_dashboard(team_id, season_id):
    """Team-specific dashboard showing stats for a specific team.

    Args:
        team_id: The team ID
        season_id: The season ID, or None for all seasons
    """
    team = db.get_our_team(team_id)
    if not team:
        st.error("Team not found")
        return

    if season_id:
        # Specific season selected
        season = db.get_season(season_id)
        if not season:
            st.error("Season not found")
            return
        st.header(f"{team.name} - {season.season_type} {season.year}")
        totals = db.get_season_totals(season_id)
        batting_stats = db.get_season_batting_stats(season_id)
        pitching_stats = db.get_season_pitching_stats(season_id)
        top_batters = db.get_top_batters(season_id, 'ba', limit=5)
        top_pitchers = db.get_top_pitchers(season_id, 'era', limit=5)
    else:
        # "All Seasons" selected - aggregate across all seasons for this team
        st.header(f"{team.name} - All Seasons")
        totals = db.get_team_all_seasons_totals(team_id)
        batting_stats = db.get_team_all_seasons_batting(team_id)
        pitching_stats = db.get_team_all_seasons_pitching(team_id)
        top_batters = db.get_team_all_seasons_top_batters(team_id, 'ba', limit=5)
        top_pitchers = db.get_team_all_seasons_top_pitchers(team_id, 'era', limit=5)

    if not totals or not totals.get('games'):
        st.info("No stats available for this selection yet.")
        return

    wins = totals.get('wins') or 0
    losses = totals.get('losses') or 0
    ties = totals.get('ties') or 0

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("Games", totals.get('games') or 0)
    with col2:
        record = f"{wins}-{losses}" + (f"-{ties}" if ties else "")
        st.metric("Record", record)
    with col3:
        win_pct = wins / (wins + losses) if (wins + losses) > 0 else 0
        st.metric("Win %", f"{win_pct:.3f}")
    with col4:
        st.metric("Runs For", totals.get('total_rf') or 0)
    with col5:
        st.metric("Runs Against", totals.get('total_ra') or 0)

    st.divider()

    # Two columns: Batting and Pitching leaders
    col_bat, col_pitch = st.columns(2)

    with col_bat:
        st.subheader("🏏 Top Batters (BA)")
        if top_batters:
            df = pd.DataFrame(top_batters)
            df = df[['player_name', 'ab', 'h', 'r', 'rbi', 'ba']]
            df.columns = ['Player', 'AB', 'H', 'R', 'RBI', 'BA']
            st.dataframe(df, width='stretch', hide_index=True)
        else:
            st.info("No batting stats yet")

    with col_pitch:
        st.subheader("⚾ Top Pitchers (ERA)")
        if top_pitchers:
            df = pd.DataFrame(top_pitchers)
            df = df[['player_name', 'ip', 'k', 'bb', 'era', 'whip']]
            df.columns = ['Player', 'IP', 'K', 'BB', 'ERA', 'WHIP']
            st.dataframe(df, width='stretch', hide_index=True)
        else:
            st.info("No pitching stats yet")

    st.divider()

    # Full batting stats
    st.subheader("📋 Full Batting Stats")
    if batting_stats:
        df = pd.DataFrame(batting_stats)
        df = df[['player_name', 'jersey_number', 'ab', 'r', 'h', 'rbi', 'bb', 'so', 'ba']]
        df.columns = ['Player', '#', 'AB', 'R', 'H', 'RBI', 'BB', 'SO', 'BA']
        st.dataframe(df, width='stretch', hide_index=True)

        total_ab = sum(s['ab'] or 0 for s in batting_stats)
        total_r = sum(s['r'] or 0 for s in batting_stats)
        total_h = sum(s['h'] or 0 for s in batting_stats)
        total_rbi = sum(s['rbi'] or 0 for s in batting_stats)
        total_bb = sum(s['bb'] or 0 for s in batting_stats)
        total_so = sum(s['so'] or 0 for s in batting_stats)
        team_ba = total_h / total_ab if total_ab > 0 else 0

        st.write(f"**Team Totals:** {total_ab} AB, {total_r} R, {total_h} H, "
                 f"{total_rbi} RBI, {total_bb} BB, {total_so} SO, {team_ba:.3f} BA")

    st.divider()

    # Full pitching stats
    st.subheader("📋 Full Pitching Stats")
    if pitching_stats:
        df = pd.DataFrame(pitching_stats)
        df = df[['player_name', 'jersey_number', 'app', 'ip', 'h', 'r', 'k', 'bb', 'era', 'whip', 'strike_pct']]
        df.columns = ['Player', '#', 'APP', 'IP', 'H', 'R', 'K', 'BB', 'ERA', 'WHIP', 'S%']
        df['S%'] = df['S%'].apply(lambda x: f"{x:.1%}" if x else "0.0%")
        st.dataframe(df, width='stretch', hide_index=True)


# =============================================================================
# GAMES PAGE
# =============================================================================

def render_games():
    """View all games for the season with add button"""
    st.title("📅 Games")

    team_id = st.session_state.get('current_team_id')
    season_id = st.session_state.get('current_season_id')

    # Determine if we can edit (specific team AND specific season required)
    can_edit = team_id is not None and season_id is not None

    if can_edit:
        # Action buttons at top (only when specific team+season selected)
        col1, col2, _ = st.columns([1, 1, 4])
        with col1:
            if st.button("➕ Add Game", width='stretch'):
                add_game_dialog()
        with col2:
            if st.button("📥 Import CSV", width='stretch'):
                import_gamechanger_dialog()
        st.divider()

    # Determine which games to show
    if season_id:
        # Specific season - show games for that season
        games = db.get_games_by_season(season_id)
        if not games:
            st.info("No games yet." + (" Click 'Add Game' or 'Import CSV' to add games." if can_edit else ""))
            return
    elif team_id:
        # Specific team but All Seasons - show all games for that team
        st.info("Showing all games across all seasons for this team.")
        games = db.get_all_games_for_team(team_id)
        if not games:
            st.info("No games found for this team.")
            return
    else:
        # All Teams - show all games for the age group
        age_group_id = st.session_state.get('current_age_group_id')
        if not age_group_id:
            st.warning("No age group selected.")
            return
        st.info("Showing all games across all teams. Select a specific team and season to add/edit games.")
        games = db.get_all_games_for_age_group(age_group_id)
        if not games:
            st.info("No games found.")
            return

    # Games table
    games_data = []
    for g in games:
        diff = g.run_differential
        diff_str = f"+{diff}" if diff and diff > 0 else str(diff) if diff else ""
        games_data.append({
            'Date': g.game_date or '',
            'Opponent': g.opponent_name or '',
            'W/L': g.win_loss or '',
            'RF': str(g.runs_for) if g.runs_for is not None else '',
            'RA': str(g.runs_against) if g.runs_against is not None else '',
            'Diff': diff_str,
            'game_id': g.id
        })

    df = pd.DataFrame(games_data)
    st.dataframe(
        df[['Date', 'Opponent', 'W/L', 'RF', 'RA', 'Diff']],
        width='stretch',
        hide_index=True
    )

    # Select game to view details
    st.divider()
    st.subheader("Game Details")

    game_options = {f"{g.game_date} vs {g.opponent_name}": g.id for g in games}
    selected_game = st.selectbox("Select a game", options=list(game_options.keys()))

    if selected_game:
        selected_game_id = game_options.get(selected_game)
        if selected_game_id:
            render_game_details(selected_game_id)


def render_game_details(game_id: int):
    """Show details for a specific game"""
    game = db.get_game(game_id)
    if not game:
        return

    # Check if editing is allowed (specific team AND specific season required)
    team_id = st.session_state.get('current_team_id')
    season_id = st.session_state.get('current_season_id')
    can_edit = team_id is not None and season_id is not None

    # Game info header with edit button
    col_info, col_edit = st.columns([4, 1])
    with col_info:
        st.write(f"**{game.game_date} vs {game.opponent_name}**")
        if game.game_time:
            st.write(f"Time: {game.game_time}")
        if game.win_loss:
            result = "Win" if game.win_loss == 'W' else "Loss" if game.win_loss == 'L' else "Tie"
            st.write(f"Result: {result} ({game.runs_for}-{game.runs_against})")
    with col_edit:
        if can_edit:
            if st.button("Edit Game", key=f"edit_game_{game_id}"):
                edit_game_dialog(game_id)

    # Batting stats for this game
    batting = db.get_batting_stats_for_game(game_id)
    if batting:
        col_bat_label, col_bat_edit = st.columns([4, 1])
        with col_bat_label:
            st.write("**Batting:**")
        with col_bat_edit:
            if can_edit:
                if st.button("Edit Stats", key=f"edit_batting_{game_id}"):
                    edit_batting_stats_dialog(game_id)

        df = pd.DataFrame(batting)
        df = df[['player_name', 'ab', 'r', 'h', 'rbi', 'bb', 'so']]
        df.columns = ['Player', 'AB', 'R', 'H', 'RBI', 'BB', 'SO']
        st.dataframe(df, width='stretch', hide_index=True)

    # Pitching stats for this game
    pitching = db.get_pitching_stats_for_game(game_id)
    if pitching:
        st.write("**Pitching:**")
        df = pd.DataFrame(pitching)
        df = df[['player_name', 'ip', 'h', 'r', 'k', 'bb', 'pitches', 'strikes']]
        df.columns = ['Player', 'IP', 'H', 'R', 'K', 'BB', 'P', 'S']
        st.dataframe(df, width='stretch', hide_index=True)


# =============================================================================
# PLAYERS PAGE
# =============================================================================

def render_players():
    """View and manage players"""
    st.title("👥 Players")

    team_id = st.session_state.get('current_team_id')

    # Only show add button when specific team is selected
    if team_id is not None:
        if st.button("➕ Add Player"):
            add_player_dialog()
        st.divider()
    else:
        st.info("Showing all players. Select a specific team to add players.")
        st.divider()

    players = db.get_all_players(active_only=False)

    if players:
        players_data = [{
            'Name': p.name,
            'Jersey #': p.jersey_number or '',
            'Bats': p.bats or '',
            'Throws': p.throws or '',
            'Active': '✓' if p.active else ''
        } for p in players]

        df = pd.DataFrame(players_data)
        st.dataframe(df, width='stretch', hide_index=True)
    else:
        st.info("No players yet." + (" Click 'Add Player' to add one." if team_id else ""))


# =============================================================================
# TEAMS PAGE (Opponent Scouting)
# =============================================================================

def render_teams():
    """View and manage opponent teams for scouting"""
    st.title("🏆 Teams (Opponent Scouting)")

    team_id = st.session_state.get('current_team_id')

    # Only show add button when specific team is selected
    if team_id is not None:
        if st.button("➕ Add Opponent Team"):
            add_opponent_team_dialog()
        st.divider()
    else:
        st.info("Showing all opponent teams. Select a specific team to add opponents.")
        st.divider()

    teams = db.get_all_teams()

    if teams:
        teams_data = []
        for t in teams:
            wins, losses, ties = db.get_record_vs_team(t.id)
            record = f"{wins}-{losses}" + (f"-{ties}" if ties else "")
            teams_data.append({
                'Team': t.name,
                'Location': t.location or '',
                'Record vs': record
            })

        df = pd.DataFrame(teams_data)
        st.dataframe(df, width='stretch', hide_index=True)
    else:
        st.info("No opponent teams yet." + (" Click 'Add Opponent Team' to add one." if team_id else ""))


# =============================================================================
# SETUP PAGE
# =============================================================================

def render_setup():
    """Setup page for managing age groups, teams, imports, and exports"""
    st.title("⚙️ Setup & Data")

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Age Groups", "Our Teams", "Import Data", "Export", "Database"
    ])

    with tab1:
        render_age_groups_setup()

    with tab2:
        render_our_teams_setup()

    with tab3:
        render_unified_import()

    with tab4:
        render_export()

    with tab5:
        render_database_management()


def render_age_groups_setup():
    """Manage age groups"""
    st.subheader("Age Groups")

    if st.button("➕ Add Age Group", key="add_ag_btn"):
        add_age_group_dialog()

    st.divider()

    age_groups = db.get_all_age_groups()

    if age_groups:
        ag_data = [{'Name': ag.name} for ag in age_groups]
        df = pd.DataFrame(ag_data)
        st.dataframe(df, width='stretch', hide_index=True)
    else:
        st.info("No age groups yet. Click 'Add Age Group' to create one.")


def render_our_teams_setup():
    """Manage our teams"""
    st.subheader("Our Teams")

    age_groups = db.get_all_age_groups()
    if not age_groups:
        st.warning("Create an age group first before adding teams.")
        return

    if st.button("➕ Add Team", key="add_team_btn"):
        add_our_team_dialog()

    st.divider()

    teams = db.get_all_our_teams(active_only=False)

    if teams:
        teams_data = []
        for t in teams:
            ag = db.get_age_group(t.age_group_id)
            teams_data.append({
                'Name': t.name,
                'Age Group': ag.name if ag else 'Unknown',
                'Location': t.location or '',
                'Active': 'Yes' if t.active else 'No'
            })
        df = pd.DataFrame(teams_data)
        st.dataframe(df, width='stretch', hide_index=True)
    else:
        st.info("No teams yet. Click 'Add Team' to create one.")


def render_unified_import():
    """Unified import for Excel and CSV files"""
    st.subheader("Import Data")

    # Check for previous import results
    if st.session_state.get('import_results'):
        results = st.session_state.import_results
        file_type = results.get('file_type', 'unknown')

        if file_type == 'excel':
            st.success("Excel import completed!")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Sheets", results.get('sheets_processed', 0))
            with col2:
                st.metric("Games", results.get('total_games', 0))
            with col3:
                st.metric("Batting", results.get('total_batting', 0))
            with col4:
                st.metric("Pitching", results.get('total_pitching', 0))
        else:
            st.success(f"CSV import completed! Created {results.get('games_created', 0)} game(s) with {results.get('total_stats', 0)} player stats.")

        if results.get('errors'):
            st.warning("Some errors occurred:")
            for err in results['errors']:
                st.write(f"- {err}")

        if st.button("Import More Files", key="clear_import"):
            del st.session_state.import_results
            st.rerun()
        return

    st.write("""
    **Supported file types:**
    - **Excel (.xlsx)**: Coach's workbook with season stats, games, batting & pitching
    - **CSV (.csv)**: GameChanger exports or standard CSV with per-game batting stats

    The importer automatically detects the file type and processes accordingly.
    """)

    # Format documentation in expandable section
    with st.expander("📖 File Format Guide & Examples", expanded=False):
        st.markdown("""
### CSV File Format

CSV files should contain **per-game batting statistics** for your team. The importer automatically detects columns based on their headers.

**Required columns:**
- **Player name** - `Player`, `Name`, or `PlayerName`
- At least one stat column

**Supported stat columns:**
| Stat | Accepted Headers |
|------|-----------------|
| At Bats | `AB`, `AtBats`, `At_Bats` |
| Runs | `R`, `Runs` |
| Hits | `H`, `Hits` |
| RBI | `RBI`, `RBIs` |
| Walks | `BB`, `Walks`, `Walk` |
| Strikeouts | `SO`, `K`, `Strikeouts` |
| Hit By Pitch | `HBP`, `HitByPitch` |
| Sacrifice | `SAC`, `Sacrifice` |

**Optional game info columns** (or enter manually during import):
| Info | Accepted Headers |
|------|-----------------|
| Date | `Date`, `GameDate`, `Game_Date` |
| Time | `Time`, `GameTime`, `Game_Time` |
| Opponent | `Opponent`, `Opp`, `Vs` |
| Result | `Result`, `W/L`, `WL`, `Win_Loss` |
| Runs For | `RF`, `RunsFor`, `Runs_For`, `Score` |
| Runs Against | `RA`, `RunsAgainst`, `Runs_Against`, `OppScore` |

---

**Example CSV format:**
```
Player,AB,R,H,RBI,BB,SO
Sarah Smith,3,2,2,1,1,0
Emma Jones,4,1,1,2,0,1
Lily Brown,3,0,1,0,0,0
Mia Davis,4,1,2,3,0,2
```

**Example with game info included:**
```
Date,Opponent,Result,RF,RA,Player,AB,R,H,RBI,BB,SO
9/6,Vipers,W,12,5,Sarah Smith,3,2,2,1,1,0
9/6,Vipers,W,12,5,Emma Jones,4,1,1,2,0,1
```

**Filename tip:** Name files like `game1_sep6_vipers.csv` to auto-detect date and opponent!

---

### Excel File Format

Excel workbooks should contain **season summary sheets** with team stats. Each sheet represents a season and should include:

- **Game Results** section with columns: Date/Opponent, W/L, RF, RA
- **Batting Stats** section with columns: #, Player, AB, R, H, RBI, BB, SO, BA
- **Pitching Stats** section with columns: #, Player, APP, IP, H, R, K, BB, HBP, P, S, S%

The importer looks for standard coach's format with headers like "GAME RESULTS", "BATTING STATS", etc.
        """)

    # Show existing data summary
    seasons = db.get_all_seasons()
    if seasons:
        with st.expander("View existing data in database"):
            for season in seasons:
                if season.id is not None:
                    summary = db.get_season_stats_summary(season.id)
                    games = summary.get('games') or 0
                    batters = summary.get('batters') or 0
                    if games > 0:
                        st.write(f"- **{season.season_type} {season.year}**: {games} games, {batters} players")

    # File uploader - accepts both Excel and CSV
    uploaded_files = st.file_uploader(
        "Drop files here or click to browse",
        type=['xlsx', 'csv'],
        accept_multiple_files=True,
        key="unified_import"
    )

    if uploaded_files:
        # Separate files by type
        excel_files = [f for f in uploaded_files if f.name.endswith('.xlsx')]
        csv_files = [f for f in uploaded_files if f.name.endswith('.csv')]

        # Initialize override variables
        override_date = ""
        override_time = ""
        override_opponent = ""
        override_wl = ""
        override_rf = ""
        override_ra = ""

        # Preview what will be imported
        if excel_files:
            st.write(f"**Excel files ({len(excel_files)}):** Team/season data will be auto-detected")
            for f in excel_files:
                st.write(f"  - {f.name}")

        if csv_files:
            st.write(f"**CSV files ({len(csv_files)}):** Game stats")

            # For CSV files, we need a team and season - show dynamic messages
            csv_team_id = st.session_state.get('current_team_id')
            csv_season_id = st.session_state.get('current_season_id')

            # Dynamic status messages based on selection state
            if not csv_team_id and not csv_season_id:
                st.error("**Step 1:** Select a **Team** from the sidebar")
                st.info("**Step 2:** Then select a **Season** to import games into")
            elif not csv_team_id:
                st.error("**Action needed:** Select a specific **Team** from the sidebar (not 'All Teams')")
            elif not csv_season_id:
                team = db.get_our_team(csv_team_id)
                team_name = team.name if team else "selected team"
                st.success(f"**Team:** {team_name}")
                st.error("**Action needed:** Select a **Season** from the sidebar (not 'All Seasons')")
            else:
                team = db.get_our_team(csv_team_id)
                season = db.get_season(csv_season_id)
                if team and season:
                    st.success(f"**Ready to import!** Games will be added to: **{team.name}** - **{season.season_type} {season.year}**")

            # Preview CSV files and detect fields
            file_previews = []
            for f in csv_files:
                # Save to temp file to preview
                with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
                    tmp.write(f.getvalue())
                    tmp_path = tmp.name

                try:
                    preview = preview_csv(tmp_path)
                    preview['file'] = f
                    preview['name'] = f.name
                    file_previews.append(preview)
                finally:
                    os.unlink(tmp_path)

            # Show what was detected
            for preview in file_previews:
                with st.expander(f"📄 {preview['name']}", expanded=len(csv_files) == 1):
                    detected = preview.get('detected_fields', {})
                    missing = preview.get('missing_fields', [])

                    if detected:
                        st.write("**Detected from file:**")
                        for field, value in detected.items():
                            st.write(f"  - {field}: `{value}`")

                    if missing:
                        st.write(f"**Not in file:** {', '.join(missing)}")

                    st.write(f"**Format:** {preview.get('format_type', 'unknown')} | **Rows:** {preview.get('row_count', 0)}")

            st.divider()
            st.write("**Game Information** (enter values or leave blank if detected above):")

            # Use detected values as defaults
            first_preview = file_previews[0] if file_previews else {}
            detected = first_preview.get('detected_fields', {})

            # Track validation errors for highlighting
            validation_errors = st.session_state.get('csv_validation_errors', [])

            # Generate time options (6 AM to 10 PM in 5-minute intervals)
            time_options = [""]
            for hour in range(6, 22):  # 6 AM to 9 PM
                for minute in range(0, 60, 5):
                    h12 = hour if hour <= 12 else hour - 12
                    if h12 == 0:
                        h12 = 12
                    ampm = "AM" if hour < 12 else "PM"
                    time_options.append(f"{h12}:{minute:02d} {ampm}")

            col1, col2 = st.columns(2)
            with col1:
                if 'game_date' in validation_errors:
                    st.markdown('<p style="color: #ff4b4b; font-weight: bold;">Game Date * (REQUIRED)</p>', unsafe_allow_html=True)

                # Date picker
                date_picker = st.date_input(
                    "Game Date *",
                    value="today",
                    key="csv_date_picker",
                    label_visibility="collapsed" if 'game_date' in validation_errors else "visible"
                )
                override_date = f"{date_picker.month}/{date_picker.day}" if date_picker else ""

            with col2:
                # Game Time - single dropdown with all options
                override_time = st.selectbox("Game Time", options=time_options, key="csv_game_time")

            col3, col4 = st.columns(2)
            with col3:
                default_opponent = detected.get('opponent') or detected.get('opponent_from_filename') or ''
                opp_label = "Opponent *"
                if 'opponent' in validation_errors:
                    st.markdown('<p style="color: #ff4b4b; font-weight: bold; margin-bottom: 0;">Opponent * (REQUIRED)</p>', unsafe_allow_html=True)
                    opp_label = "Opponent *"
                override_opponent = st.text_input(
                    opp_label,
                    value=default_opponent,
                    placeholder="e.g., NJ Vipers",
                    key="csv_opp_override",
                    help="Required. Enter opponent team name",
                    label_visibility="collapsed" if 'opponent' in validation_errors else "visible"
                )
            with col4:
                default_wl = detected.get('win_loss') or ''
                override_wl = st.selectbox(
                    "Result",
                    options=['', 'W', 'L', 'T'],
                    index=['', 'W', 'L', 'T'].index(default_wl) if default_wl in ['', 'W', 'L', 'T'] else 0,
                    key="csv_wl_override",
                    help="Optional. Win/Loss/Tie"
                )

            col5, col6 = st.columns(2)
            with col5:
                default_rf = detected.get('runs_for') or ''
                override_rf = st.text_input(
                    "Our Score (RF)",
                    value=str(default_rf) if default_rf else '',
                    placeholder="e.g., 12",
                    key="csv_rf_override",
                    help="Optional. Runs For"
                )
            with col6:
                default_ra = detected.get('runs_against') or ''
                override_ra = st.text_input(
                    "Their Score (RA)",
                    value=str(default_ra) if default_ra else '',
                    placeholder="e.g., 5",
                    key="csv_ra_override",
                    help="Optional. Runs Against"
                )

            # Clear validation errors after displaying
            if 'csv_validation_errors' in st.session_state:
                del st.session_state['csv_validation_errors']

        st.divider()

        # Import button with validation
        can_import = True
        missing_requirements = []

        # Check team and season selection for CSV files
        if csv_files:
            if not st.session_state.get('current_team_id'):
                can_import = False
                missing_requirements.append("Select a **Team** from the sidebar")
            if not st.session_state.get('current_season_id'):
                can_import = False
                missing_requirements.append("Select a **Season** from the sidebar")

            # Check required fields only if team/season are selected
            if st.session_state.get('current_season_id'):
                if not override_date:
                    can_import = False
                    missing_requirements.append("Enter a **Game Date**")
                if not override_opponent:
                    can_import = False
                    missing_requirements.append("Enter an **Opponent** name")

        # Show what's missing
        if missing_requirements:
            st.warning("**Before importing, please:**")
            for req in missing_requirements:
                st.markdown(f"- {req}")

        # Button - always visible but validates on click
        import_clicked = st.button(
            "Import All Files",
            type="primary",
            key="import_all_btn",
            disabled=not (excel_files or csv_files)  # Only disable if no files
        )

        if import_clicked:
            # Validate CSV requirements
            if csv_files:
                validation_errors = []
                if not st.session_state.get('current_team_id'):
                    st.error("Please select a Team from the sidebar first!")
                    st.stop()
                if not st.session_state.get('current_season_id'):
                    st.error("Please select a Season from the sidebar first!")
                    st.stop()
                if not override_date:
                    validation_errors.append('game_date')
                if not override_opponent:
                    validation_errors.append('opponent')

                if validation_errors:
                    st.session_state['csv_validation_errors'] = validation_errors
                    st.error("Please fill in all required fields (marked with *) and try again.")
                    st.rerun()

        if import_clicked and can_import:
            total_results = {
                'file_type': 'mixed' if (excel_files and csv_files) else ('excel' if excel_files else 'csv'),
                'sheets_processed': 0,
                'total_games': 0,
                'total_batting': 0,
                'total_pitching': 0,
                'games_created': 0,
                'total_stats': 0,
                'errors': []
            }

            progress = st.progress(0)
            status = st.empty()
            total_files = len(excel_files) + len(csv_files)
            processed = 0

            # Process Excel files first
            for f in excel_files:
                status.text(f"Importing {f.name}...")

                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(f.getvalue())
                    tmp_path = tmp.name

                try:
                    result = import_file(tmp_path, verbose=False)

                    total_results['sheets_processed'] += result.get('sheets_processed', 0)
                    total_results['total_games'] += result.get('total_games', 0)
                    total_results['total_batting'] += result.get('total_batting', 0)
                    total_results['total_pitching'] += result.get('total_pitching', 0)

                    if result.get('errors'):
                        total_results['errors'].extend(result['errors'])

                except Exception as e:
                    total_results['errors'].append(f"{f.name}: {e}")

                finally:
                    os.unlink(tmp_path)

                processed += 1
                progress.progress(processed / total_files)

            # Process CSV files
            csv_season_id = st.session_state.get('current_season_id')
            for f in csv_files:
                status.text(f"Importing {f.name}...")

                # Use override values, falling back to auto-detection in import_file
                game_date = override_date if override_date else None
                game_time = override_time if override_time else None
                opponent = override_opponent if override_opponent else None
                win_loss = override_wl if override_wl else None
                runs_for = int(override_rf) if override_rf and override_rf.isdigit() else None
                runs_against = int(override_ra) if override_ra and override_ra.isdigit() else None

                with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
                    tmp.write(f.getvalue())
                    tmp_path = tmp.name

                try:
                    result = import_file(
                        tmp_path, season_id=csv_season_id,
                        game_date=game_date, game_time=game_time,
                        opponent=opponent, win_loss=win_loss,
                        runs_for=runs_for, runs_against=runs_against,
                        verbose=False
                    )

                    if result.get('game_id'):
                        total_results['games_created'] += 1
                        total_results['total_stats'] += result.get('stats_imported', 0)

                    if result.get('errors'):
                        total_results['errors'].extend(result['errors'])

                except Exception as e:
                    total_results['errors'].append(f"{f.name}: {e}")

                finally:
                    os.unlink(tmp_path)

                processed += 1
                progress.progress(processed / total_files)

            status.empty()
            progress.empty()

            st.session_state.import_results = total_results
            st.rerun()


def render_export():
    """Export data to various formats"""
    st.subheader("Export Data")

    if not st.session_state.get('current_season_id'):
        st.warning("Please select a season from the sidebar to export.")
        return

    season_id = st.session_state.current_season_id
    season = db.get_season(season_id)
    if not season:
        st.error("Selected season not found")
        return

    st.write(f"Export data for: **{season.season_type} {season.year}**")

    st.write("---")

    st.write("**Export to Excel**")
    st.write("Generate an Excel workbook in the coach's format with games, batting, and pitching stats.")

    if st.button("Generate Excel Export", type="primary"):
        with st.spinner("Generating export..."):
            excel_data = generate_excel_export(season_id)

        filename = f"{season.season_type}_{season.year}_stats.xlsx"
        st.download_button(
            label="📥 Download Excel File",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


def render_database_management():
    """Database management - view stats and reset"""
    st.subheader("Database Management")

    # Show current database stats
    stats = db.get_database_stats()

    st.write("**Current Database Contents:**")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Age Groups", stats['age_groups'])
        st.metric("Teams", stats['teams'])
    with col2:
        st.metric("Seasons", stats['seasons'])
        st.metric("Games", stats['games'])
    with col3:
        st.metric("Players", stats['players'])
    with col4:
        st.metric("Batting Records", stats['batting_stats'])
        st.metric("Pitching Records", stats['pitching_stats'])

    st.divider()

    # Reset database section
    st.write("**Reset Database**")
    st.warning("This will permanently delete ALL data including teams, seasons, games, and player stats. This cannot be undone!")

    # Two-step confirmation
    if st.session_state.get('confirm_reset_step1'):
        st.error("Are you absolutely sure? Type 'RESET' below and click confirm to proceed.")
        confirm_text = st.text_input("Type RESET to confirm:", key="reset_confirm_text")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Confirm Reset", type="primary"):
                if confirm_text == "RESET":
                    db.reset_database()
                    st.session_state.confirm_reset_step1 = False
                    st.success("Database has been reset successfully!")
                    st.rerun()
                else:
                    st.error("You must type 'RESET' exactly to confirm.")
        with col2:
            if st.button("Cancel"):
                st.session_state.confirm_reset_step1 = False
                st.rerun()
    else:
        if st.button("Reset Database", type="secondary"):
            st.session_state.confirm_reset_step1 = True
            st.rerun()


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def generate_excel_export(season_id: int) -> bytes:
    """Generate Excel export in coach's format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active

    season = db.get_season(season_id)
    ws.title = season.name

    header_font = Font(bold=True, size=11)
    subheader_font = Font(bold=True, size=10)
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    ws['A1'] = f"TEAM NAME: PA Chaos 12U Taranto"
    ws['A1'].font = header_font

    wins, losses, ties = db.get_season_record(season_id)
    record = f"{wins}-{losses}" + (f"-{ties}" if ties else "")
    ws['A6'] = f"{season.season_type.upper()} {season.year} RECORD: {record}"
    ws['A6'].font = header_font

    ws['A8'] = "GAME RESULTS"
    ws['G8'] = "PITCHING STATS"
    ws['Y8'] = "BATTING STATS"

    game_headers = ['DATE/OPP', 'W/L', 'RF', 'RA', 'DIFF']
    for col, header in enumerate(game_headers, start=1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = subheader_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    pitch_headers = ['#', 'PLAYER', 'APP', 'IP', 'H', 'R', 'K', 'BB', 'HBP', 'P', 'S', 'S%']
    for col_offset, header in enumerate(pitch_headers):
        cell = ws.cell(row=9, column=7+col_offset, value=header)
        cell.font = subheader_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    bat_headers = ['#', 'PLAYER', 'AB', 'R', 'H', 'RBI', 'BB', 'SO', 'BA']
    for col_offset, header in enumerate(bat_headers):
        cell = ws.cell(row=9, column=25+col_offset, value=header)
        cell.font = subheader_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    games = db.get_games_by_season(season_id)
    batting_stats = db.get_season_batting_stats(season_id)
    pitching_stats = db.get_season_pitching_stats(season_id)

    max_rows = max(len(games), len(batting_stats), len(pitching_stats))

    for row_idx in range(max_rows):
        row = 10 + row_idx

        if row_idx < len(games):
            g = games[row_idx]
            ws.cell(row=row, column=1, value=f"{g.game_date} vs {g.opponent_name}")
            ws.cell(row=row, column=2, value=g.win_loss or '')
            ws.cell(row=row, column=3, value=g.runs_for or '')
            ws.cell(row=row, column=4, value=g.runs_against or '')
            diff = g.run_differential
            ws.cell(row=row, column=5, value=diff if diff else '')

        if row_idx < len(pitching_stats):
            p = pitching_stats[row_idx]
            ws.cell(row=row, column=7, value=p.get('jersey_number', ''))
            ws.cell(row=row, column=8, value=p['player_name'])
            ws.cell(row=row, column=9, value=p.get('app', 0))
            ws.cell(row=row, column=10, value=p.get('ip', 0))
            ws.cell(row=row, column=11, value=p.get('h', 0))
            ws.cell(row=row, column=12, value=p.get('r', 0))
            ws.cell(row=row, column=13, value=p.get('k', 0))
            ws.cell(row=row, column=14, value=p.get('bb', 0))
            ws.cell(row=row, column=15, value=p.get('hbp', 0))
            ws.cell(row=row, column=16, value=p.get('pitches', 0))
            ws.cell(row=row, column=17, value=p.get('strikes', 0))
            ws.cell(row=row, column=18, value=p.get('strike_pct', 0))

        if row_idx < len(batting_stats):
            b = batting_stats[row_idx]
            ws.cell(row=row, column=25, value=b.get('jersey_number', ''))
            ws.cell(row=row, column=26, value=b['player_name'])
            ws.cell(row=row, column=27, value=b.get('ab', 0))
            ws.cell(row=row, column=28, value=b.get('r', 0))
            ws.cell(row=row, column=29, value=b.get('h', 0))
            ws.cell(row=row, column=30, value=b.get('rbi', 0))
            ws.cell(row=row, column=31, value=b.get('bb', 0))
            ws.cell(row=row, column=32, value=b.get('so', 0))
            ws.cell(row=row, column=33, value=b.get('ba', 0))

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['Z'].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main application entry point"""
    # Check password first
    if not check_password():
        return

    db.init_database()

    page = render_sidebar()

    if page == "Dashboard":
        render_dashboard()
    elif page == "Games":
        render_games()
    elif page == "Players":
        render_players()
    elif page == "Teams":
        render_teams()
    elif page == "Setup":
        render_setup()


if __name__ == "__main__":
    main()