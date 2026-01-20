"""
Complete structure analysis of 12U Teams.xlsx - finding all season blocks
"""
import openpyxl
from openpyxl.utils import get_column_letter

def find_all_sections(ws):
    """Find all section markers (RECORD lines, GAME RESULTS headers, etc.)"""
    sections = []
    for row_num in range(1, ws.max_row + 1):
        for col_num in range(1, min(50, ws.max_column + 1)):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                if 'RECORD:' in val:
                    sections.append({'type': 'RECORD', 'row': row_num, 'col': col_num, 'value': val})
                elif 'GAME RESULTS' in val:
                    sections.append({'type': 'GAME_RESULTS', 'row': row_num, 'col': col_num, 'value': val})
                elif 'PITCHING -' in val:
                    sections.append({'type': 'PITCHING', 'row': row_num, 'col': col_num, 'value': val})
                elif 'BATTING -' in val:
                    sections.append({'type': 'BATTING', 'row': row_num, 'col': col_num, 'value': val})
                elif val.startswith('TEAM NAME:'):
                    sections.append({'type': 'TEAM', 'row': row_num, 'col': col_num, 'value': val})
    return sections

def get_column_headers_at_row(ws, row_num, start_col=1, end_col=None):
    """Get all column headers at a specific row"""
    if end_col is None:
        end_col = ws.max_column
    headers = {}
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            headers[get_column_letter(col)] = cell.value
    return headers

def count_game_rows(ws, start_row, max_rows=100):
    """Count rows that look like game data"""
    count = 0
    for row_num in range(start_row, min(start_row + max_rows, ws.max_row + 1)):
        cell = ws.cell(row=row_num, column=1)
        if cell.value:
            val = str(cell.value)
            # Game entries typically have date pattern like "9/6 vs."
            if '/' in val and 'vs.' in val:
                count += 1
            elif val.startswith(('TEAM NAME:', 'FALL', 'SPRING', 'SUMMER')):
                break
    return count

def analyze_complete(filepath):
    """Complete analysis of the workbook structure"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    print("=" * 120)
    print("COMPLETE WORKBOOK STRUCTURE ANALYSIS: 12U Teams.xlsx")
    print("=" * 120)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n\n{'#'*120}")
        print(f"SHEET: '{sheet_name}'")
        print(f"{'#'*120}")
        print(f"Total dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column} ({get_column_letter(ws.max_column)})")

        # Find all sections
        sections = find_all_sections(ws)

        # Group by type
        teams = [s for s in sections if s['type'] == 'TEAM']
        records = [s for s in sections if s['type'] == 'RECORD']
        game_results = [s for s in sections if s['type'] == 'GAME_RESULTS']
        pitching = [s for s in sections if s['type'] == 'PITCHING']
        batting = [s for s in sections if s['type'] == 'BATTING']

        print(f"\n--- STRUCTURE SUMMARY ---")
        print(f"  Teams found: {len(teams)}")
        print(f"  Season records found: {len(records)}")
        print(f"  Game results sections: {len(game_results)}")
        print(f"  Pitching sections: {len(pitching)}")
        print(f"  Batting sections: {len(batting)}")

        print(f"\n--- TEAMS ---")
        for t in teams:
            print(f"  Row {t['row']}: {t['value']}")

        print(f"\n--- SEASON RECORDS ---")
        for r in records:
            print(f"  Row {r['row']}: {r['value']}")

        print(f"\n--- GAME RESULTS SECTIONS ---")
        for gr in game_results:
            print(f"\n  ROW {gr['row']}: {gr['value']}")
            # Get column structure from this row
            headers = get_column_headers_at_row(ws, gr['row'])
            print(f"    Column headers at row {gr['row']}:")
            for col, val in sorted(headers.items()):
                print(f"      {col}: {val}")

            # Get detailed headers from next row
            next_headers = get_column_headers_at_row(ws, gr['row'] + 1)
            if next_headers:
                print(f"    Column headers at row {gr['row'] + 1} (detail headers):")
                for col, val in sorted(next_headers.items()):
                    print(f"      {col}: {val}")

            # Count games
            game_count = count_game_rows(ws, gr['row'] + 2)
            print(f"    Number of game entries: {game_count}")

        # Analyze pitching column structure
        print(f"\n--- PITCHING COLUMN STRUCTURE ---")
        for p in pitching[:3]:  # Just show first 3
            print(f"\n  '{p['value']}' at row {p['row']}, column {get_column_letter(p['col'])}")
            # Get headers from the next row
            headers = get_column_headers_at_row(ws, p['row'] + 1, p['col'], p['col'] + 20)
            for col, val in sorted(headers.items()):
                print(f"    {col}: {val}")

        # Analyze batting column structure
        print(f"\n--- BATTING COLUMN STRUCTURE ---")
        for b in batting[:3]:  # Just show first 3
            print(f"\n  '{b['value']}' at row {b['row']}, column {get_column_letter(b['col'])}")
            # Get headers from the next row
            headers = get_column_headers_at_row(ws, b['row'] + 1, b['col'], b['col'] + 15)
            for col, val in sorted(headers.items()):
                print(f"    {col}: {val}")

        # Sample actual pitching data
        if pitching:
            print(f"\n--- SAMPLE PITCHING DATA ---")
            p = pitching[0]
            header_row = p['row'] + 1
            for data_row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
                row_data = []
                for col in range(p['col'], min(p['col'] + 18, ws.max_column + 1)):
                    cell = ws.cell(row=data_row, column=col)
                    if cell.value is not None:
                        row_data.append(f"{get_column_letter(col)}:{cell.value}")
                if row_data:
                    print(f"  Row {data_row}: {row_data}")

        # Sample actual batting data
        if batting:
            print(f"\n--- SAMPLE BATTING DATA ---")
            b = batting[0]
            header_row = b['row'] + 1
            for data_row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
                row_data = []
                for col in range(b['col'], min(b['col'] + 10, ws.max_column + 1)):
                    cell = ws.cell(row=data_row, column=col)
                    if cell.value is not None:
                        row_data.append(f"{get_column_letter(col)}:{cell.value}")
                if row_data:
                    print(f"  Row {data_row}: {row_data}")

    wb.close()
    print(f"\n\n{'='*120}")
    print("ANALYSIS COMPLETE")
    print(f"{'='*120}")

if __name__ == "__main__":
    filepath = r"c:\Wolvryn Projects\teamstats\test-data\12U Teams.xlsx"
    analyze_complete(filepath)