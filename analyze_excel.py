"""
Script to analyze the structure of the 12U Teams.xlsx Excel file
"""
import openpyxl
from openpyxl.utils import get_column_letter

def analyze_workbook(filepath):
    """Analyze the structure of an Excel workbook"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    print("=" * 80)
    print(f"WORKBOOK ANALYSIS: {filepath}")
    print(f"Number of sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")
    print("=" * 80)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'#' * 80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'#' * 80}")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

        # Analyze header rows (rows 1-10)
        print(f"\n--- HEADER ROWS (1-10) ---")
        for row_num in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    row_data.append(f"{get_column_letter(col_num)}{row_num}: {repr(cell.value)}")
            if row_data:
                print(f"Row {row_num}: {row_data}")
            else:
                print(f"Row {row_num}: [empty]")

        # Find potential column headers (look for rows with multiple text values)
        print(f"\n--- POTENTIAL COLUMN HEADERS ---")
        for row_num in range(1, min(15, ws.max_row + 1)):
            text_cells = []
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None and isinstance(cell.value, str):
                    text_cells.append((get_column_letter(col_num), cell.value))
            if len(text_cells) >= 3:  # Likely a header row if 3+ text values
                print(f"Row {row_num} (likely header): {text_cells}")

        # Sample data rows (first few after headers)
        print(f"\n--- SAMPLE DATA ROWS (11-20) ---")
        for row_num in range(11, min(21, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    row_data.append(f"{get_column_letter(col_num)}: {repr(cell.value)}")
            if row_data:
                print(f"Row {row_num}: {row_data}")

        # Identify stat categories
        print(f"\n--- STAT CATEGORIES DETECTED ---")
        all_headers = set()
        for row_num in range(1, min(15, ws.max_row + 1)):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None and isinstance(cell.value, str):
                    all_headers.add(cell.value.strip())

        # Common baseball stat abbreviations
        batting_stats = ['AB', 'H', 'R', 'RBI', 'HR', 'BB', 'SO', 'K', 'AVG', '1B', '2B', '3B',
                        'SB', 'CS', 'OBP', 'SLG', 'OPS', 'HBP', 'SAC', 'SF', 'PA', 'TB']
        pitching_stats = ['W', 'L', 'ERA', 'IP', 'ER', 'H', 'BB', 'SO', 'K', 'WHIP', 'SV',
                         'HLD', 'BS', 'CG', 'SHO', 'BF', 'HR', 'WP', 'HBP']
        team_stats = ['W', 'L', 'T', 'RF', 'RA', 'PCT', 'GB', 'STRK', 'RS', 'Wins', 'Losses']

        found_batting = [s for s in all_headers if s.upper() in [x.upper() for x in batting_stats]]
        found_pitching = [s for s in all_headers if s.upper() in [x.upper() for x in pitching_stats]]
        found_team = [s for s in all_headers if s.upper() in [x.upper() for x in team_stats]]

        print(f"All unique header values: {sorted(all_headers)}")
        if found_batting:
            print(f"Batting stats found: {found_batting}")
        if found_pitching:
            print(f"Pitching stats found: {found_pitching}")
        if found_team:
            print(f"Team/Record stats found: {found_team}")

        # Check for merged cells
        if ws.merged_cells.ranges:
            print(f"\n--- MERGED CELLS ---")
            for merged_range in ws.merged_cells.ranges:
                print(f"  {merged_range}")

        # Data organization pattern
        print(f"\n--- DATA ORGANIZATION ---")
        # Check if data is organized by team (look for team name patterns)
        team_indicators = []
        for row_num in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_num, column=1)
            if cell_a.value and isinstance(cell_a.value, str):
                if len(cell_a.value) > 3 and not cell_a.value.isupper():
                    team_indicators.append((row_num, cell_a.value))

        if team_indicators[:10]:
            print(f"Potential team/player names in column A: {team_indicators[:10]}")

        print("\n")

    wb.close()
    return wb.sheetnames

if __name__ == "__main__":
    filepath = r"c:\Wolvryn Projects\teamstats\test-data\12U Teams.xlsx"
    analyze_workbook(filepath)