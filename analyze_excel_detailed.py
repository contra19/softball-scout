"""
Detailed structure analysis of 12U Teams.xlsx
"""
import openpyxl
from openpyxl.utils import get_column_letter

def find_team_blocks(ws):
    """Find where each team's data block starts"""
    team_blocks = []
    for row_num in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=1)
        if cell.value and isinstance(cell.value, str):
            if cell.value.startswith('TEAM NAME:'):
                team_blocks.append((row_num, cell.value))
    return team_blocks

def analyze_team_block(ws, start_row, end_row=None):
    """Analyze a single team's data block"""
    if end_row is None:
        end_row = start_row + 60  # Analyze next 60 rows

    team_info = {
        'name': None,
        'location': None,
        'record': None,
        'game_results_header_row': None,
        'pitching_header_row': None,
        'batting_header_row': None,
        'pitching_columns': {},
        'batting_columns': {},
        'game_results_columns': {},
        'game_count': 0,
        'pitcher_count': 0,
        'batter_count': 0
    }

    for row_num in range(start_row, min(end_row, ws.max_row + 1)):
        cell_a = ws.cell(row=row_num, column=1)
        if cell_a.value:
            val = str(cell_a.value)
            if val.startswith('TEAM NAME:'):
                team_info['name'] = val.replace('TEAM NAME:', '').strip()
            elif val.startswith('LOCATION:'):
                team_info['location'] = val.replace('LOCATION:', '').strip()
            elif 'RECORD:' in val:
                team_info['record'] = val
            elif 'GAME RESULTS' in val:
                team_info['game_results_header_row'] = row_num
                # Get column headers from the next row
                for col in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=row_num, column=col)
                    if header_cell.value:
                        team_info['game_results_columns'][get_column_letter(col)] = header_cell.value

        # Check for pitching section header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col)
            if cell.value and isinstance(cell.value, str):
                if 'PITCHING' in cell.value and team_info['pitching_header_row'] is None:
                    team_info['pitching_header_row'] = row_num
                    # Get pitching column headers from the row below
                    next_row = row_num + 1
                    for c in range(col, col + 20):  # Check next 20 columns
                        if c <= ws.max_column:
                            h = ws.cell(row=next_row, column=c)
                            if h.value:
                                team_info['pitching_columns'][get_column_letter(c)] = h.value
                elif 'BATTING' in cell.value and team_info['batting_header_row'] is None:
                    team_info['batting_header_row'] = row_num
                    # Get batting column headers from the row below
                    next_row = row_num + 1
                    for c in range(col, col + 15):  # Check next 15 columns
                        if c <= ws.max_column:
                            h = ws.cell(row=next_row, column=c)
                            if h.value:
                                team_info['batting_columns'][get_column_letter(c)] = h.value

    return team_info

def count_data_rows(ws, header_row, end_row):
    """Count how many data rows exist after a header row"""
    count = 0
    for row_num in range(header_row + 2, min(end_row, ws.max_row + 1)):
        cell = ws.cell(row=row_num, column=1)
        if cell.value and not str(cell.value).startswith('TEAM NAME:'):
            # Check if this looks like game data (date pattern)
            val = str(cell.value)
            if '/' in val and 'vs.' in val:
                count += 1
            elif val.startswith(('FALL', 'SPRING', 'SUMMER')):
                break
    return count

def detailed_analysis(filepath):
    """Perform detailed analysis of the workbook"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    print("=" * 100)
    print("DETAILED WORKBOOK STRUCTURE ANALYSIS")
    print("=" * 100)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*100}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*100}")

        # Find all team blocks
        team_blocks = find_team_blocks(ws)
        print(f"\nNumber of teams in sheet: {len(team_blocks)}")

        for i, (start_row, team_label) in enumerate(team_blocks):
            # Determine end row (start of next team or end of sheet)
            if i + 1 < len(team_blocks):
                end_row = team_blocks[i + 1][0]
            else:
                end_row = ws.max_row

            print(f"\n{'-'*80}")
            print(f"TEAM {i+1}: {team_label}")
            print(f"Data starts at row: {start_row}, ends before row: {end_row}")
            print(f"{'-'*80}")

            team_info = analyze_team_block(ws, start_row, end_row)

            print(f"  Team Name: {team_info['name']}")
            print(f"  Location: {team_info['location']}")
            print(f"  Record: {team_info['record']}")

            if team_info['game_results_columns']:
                print(f"\n  GAME RESULTS SECTION (Row {team_info['game_results_header_row']}):")
                for col, header in sorted(team_info['game_results_columns'].items()):
                    print(f"    Column {col}: {header}")

            if team_info['pitching_columns']:
                print(f"\n  PITCHING SECTION (Row {team_info['pitching_header_row']}):")
                for col, header in sorted(team_info['pitching_columns'].items()):
                    print(f"    Column {col}: {header}")

            if team_info['batting_columns']:
                print(f"\n  BATTING SECTION (Row {team_info['batting_header_row']}):")
                for col, header in sorted(team_info['batting_columns'].items()):
                    print(f"    Column {col}: {header}")

            # Sample some actual data
            print(f"\n  SAMPLE DATA (first 5 rows after headers):")
            data_start = (team_info['game_results_header_row'] or start_row) + 2
            for row_num in range(data_start, min(data_start + 5, end_row)):
                row_preview = []
                for col_num in range(1, min(35, ws.max_column + 1)):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        row_preview.append(f"{get_column_letter(col_num)}:{repr(cell.value)[:30]}")
                if row_preview:
                    print(f"    Row {row_num}: {', '.join(row_preview[:10])}...")

    wb.close()

if __name__ == "__main__":
    filepath = r"c:\Wolvryn Projects\teamstats\test-data\12U Teams.xlsx"
    detailed_analysis(filepath)