"""
Import tool for coach's existing Excel data
Parses the 3-panel layout and imports into SQLite database
"""

from openpyxl import load_workbook
from typing import List, Dict, Tuple, Optional
import re
from database import (
    get_or_create_season, get_or_create_player, create_game,
    add_batting_stats, add_pitching_stats, get_db,
    get_or_create_age_group, get_or_create_our_team
)


def parse_season_from_text(text: str) -> Tuple[Optional[str], Optional[int]]:
    """
    Parse season type and year from text like "FALL 2025 RECORD: 30-9"
    Returns (season_type, year) or (None, None)
    """
    if not text:
        return None, None

    match = re.search(r'(FALL|SPRING)\s+(\d{4})', text.upper())
    if match:
        return match.group(1).capitalize(), int(match.group(2))
    return None, None


def parse_game_entry(text: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse game date and opponent from text like "9/6 vs. NJ Vipers 12U White"
    Returns (date, opponent) or (None, None)
    """
    if not text:
        return None, None

    # Pattern: date vs. opponent or date vs opponent
    match = re.match(r'^(\d{1,2}/\d{1,2})\s+vs\.?\s+(.+)$', str(text).strip())
    if match:
        return match.group(1), match.group(2).strip()
    return None, None


def parse_record(text: str) -> Tuple[int, int, int]:
    """
    Parse W-L or W-L-T record from text like "30-9" or "36-15-1"
    Returns (wins, losses, ties)
    """
    if not text:
        return 0, 0, 0

    # Match patterns like "30-9" or "36-15-1"
    match = re.search(r'(\d+)-(\d+)(?:-(\d+))?', str(text))
    if match:
        wins = int(match.group(1))
        losses = int(match.group(2))
        ties = int(match.group(3)) if match.group(3) else 0
        return wins, losses, ties
    return 0, 0, 0


def clean_player_name(name: str) -> str:
    """Remove handedness indicators and clean player name. Returns empty for headers."""
    if not name:
        return ""
    name = str(name).strip()
    # Skip header row values
    if name.upper() in ['PLAYER', 'NAME', 'PLAYERS', '#', '']:
        return ""
    # Remove (R), (L), (S) suffix
    name = re.sub(r'\s*\([RLS]\)\s*$', '', name)
    return name.strip()


def safe_int(val) -> int:
    """Safely convert value to int"""
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def format_jersey(val) -> Optional[str]:
    """Format jersey number as whole number string, or None if invalid"""
    if val is None or val == '':
        return None
    # Skip header row values
    str_val = str(val).strip()
    if str_val.upper() in ['#', 'NO', 'NO.', 'NUM', 'NUMBER', 'JERSEY']:
        return None
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return str_val if str_val else None


def is_header_or_invalid_player(name: str, jersey: any) -> bool:
    """Check if a row appears to be a header row rather than actual player data"""
    if not name:
        return True
    name_upper = str(name).strip().upper()
    # Common header values for player name column
    if name_upper in ['PLAYER', 'NAME', 'PLAYERS', 'TOTALS', '#', '']:
        return True
    # Check if jersey looks like a header
    if jersey is not None:
        jersey_str = str(jersey).strip().upper()
        if jersey_str in ['#', 'NO', 'NO.', 'NUM', 'NUMBER', 'JERSEY']:
            return True
    return False


def safe_float(val) -> float:
    """Safely convert value to float"""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def extract_team_info(ws) -> Dict:
    """
    Extract team info from worksheet header rows.
    Returns dict with: team_name, age_group, location
    """
    team_info = {
        'team_name': None,
        'age_group': None,
        'location': None,
        'full_name': None
    }

    # Look in first 10 rows for team info
    for row in range(1, 10):
        cell_val = ws.cell(row=row, column=1).value
        if not cell_val:
            continue
        cell_str = str(cell_val).strip()

        # Team name: "TEAM NAME: PA Chaos 12U Taranto"
        if 'TEAM NAME:' in cell_str.upper():
            full_name = cell_str.split(':', 1)[1].strip()
            team_info['full_name'] = full_name
            team_info['team_name'] = full_name

            # Extract age group (8U, 10U, 11U, 12U, 14U, 16U, 18U)
            age_match = re.search(r'(\d{1,2}U)', full_name, re.IGNORECASE)
            if age_match:
                team_info['age_group'] = age_match.group(1).upper()

        # Location: "LOCATION: Garnett Valley, PA"
        elif 'LOCATION:' in cell_str.upper():
            team_info['location'] = cell_str.split(':', 1)[1].strip()

    return team_info


def get_or_create_team_from_info(team_info: Dict, verbose: bool = True) -> Optional[int]:
    """
    Create age group and team from extracted info.
    Returns team_id or None if not enough info.
    """
    if not team_info.get('age_group') or not team_info.get('team_name'):
        return None

    # Create age group
    age_group = team_info['age_group']
    # Sort order: 8U=8, 10U=10, 12U=12, etc.
    sort_order = int(age_group.replace('U', ''))
    age_group_id = get_or_create_age_group(age_group, sort_order)

    if verbose:
        print(f"Age group: {age_group} (ID: {age_group_id})")

    # Create team
    team_id = get_or_create_our_team(
        name=team_info['team_name'],
        age_group_id=age_group_id,
        location=team_info.get('location')
    )

    if verbose:
        print(f"Team: {team_info['team_name']} (ID: {team_id})")

    return team_id


def import_fall_sheet(ws, team_id: int = None, verbose: bool = True) -> Dict:
    """
    Import the 'Fall' sheet format

    Layout:
    - Row 1-5: Team info
    - Row 7: Season record
    - Row 8: Section headers (Game Results | Pitching Stats | Batting Stats)
    - Row 9: Column headers
    - Row 10+: Data

    Columns:
    - A: Game entry (date vs. opponent)
    - B: W/L
    - C: RF
    - D: RA
    - E: DIFF
    - G-W: Pitching stats
    - Y-AG: Batting stats
    """
    results = {
        'season_id': None,
        'games_imported': 0,
        'batting_records': 0,
        'pitching_records': 0,
        'errors': []
    }

    # Get season from row 7
    season_text = ws['A7'].value or ws['B7'].value
    season_type, year = parse_season_from_text(str(season_text) if season_text else "")

    if not season_type or not year:
        # Try to find it elsewhere
        for row in range(1, 10):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val:
                season_type, year = parse_season_from_text(str(cell_val))
                if season_type and year:
                    break

    if not season_type or not year:
        results['errors'].append("Could not determine season from sheet")
        return results

    season_id = get_or_create_season(year, season_type, team_id)
    results['season_id'] = season_id

    if verbose:
        print(f"Importing {season_type} {year} (Season ID: {season_id})")

    # First pass: collect all unique players from batting and pitching sections
    # Batting is in columns Y-AG (25-33), players in column Z (26)
    # Pitching is in columns G-W (7-23), players in column H (8)

    batting_players = {}  # name -> {stats}
    pitching_players = {}  # name -> {stats}
    games = []  # [(date, opponent, w_l, rf, ra)]

    # Read data starting from row 9 (first game row after headers on row 8)
    row = 9
    while row < 500:  # Safety limit
        # Check if we hit an empty row or another section
        game_entry = ws.cell(row=row, column=1).value

        # Read game info
        game_date, opponent = parse_game_entry(game_entry)
        if game_date and opponent:
            w_l = ws.cell(row=row, column=2).value
            rf = safe_int(ws.cell(row=row, column=3).value)
            ra = safe_int(ws.cell(row=row, column=4).value)
            games.append((game_date, opponent, w_l, rf, ra))

        # Read pitching stats (columns G-W, player in H)
        pitcher_name_raw = ws.cell(row=row, column=8).value
        pitcher_jersey_raw = ws.cell(row=row, column=7).value
        pitcher_name = clean_player_name(pitcher_name_raw)
        if pitcher_name and not is_header_or_invalid_player(pitcher_name_raw, pitcher_jersey_raw):
            pitching_players[pitcher_name] = {
                'jersey': pitcher_jersey_raw,
                'app': safe_int(ws.cell(row=row, column=9).value),
                'ip': safe_float(ws.cell(row=row, column=10).value),
                'h': safe_int(ws.cell(row=row, column=11).value),
                'r': safe_int(ws.cell(row=row, column=13).value),
                'k': safe_int(ws.cell(row=row, column=15).value),
                'bb': safe_int(ws.cell(row=row, column=17).value),
                'hbp': safe_int(ws.cell(row=row, column=20).value),
                'pitches': safe_int(ws.cell(row=row, column=21).value),
                'strikes': safe_int(ws.cell(row=row, column=22).value),
            }

        # Read batting stats (columns 25-33, player in column 26)
        batter_name_raw = ws.cell(row=row, column=26).value
        batter_jersey_raw = ws.cell(row=row, column=25).value
        batter_name = clean_player_name(batter_name_raw)
        if batter_name and not is_header_or_invalid_player(batter_name_raw, batter_jersey_raw):
            batting_players[batter_name] = {
                'jersey': batter_jersey_raw,
                'ab': safe_int(ws.cell(row=row, column=27).value),
                'r': safe_int(ws.cell(row=row, column=28).value),
                'h': safe_int(ws.cell(row=row, column=29).value),
                'rbi': safe_int(ws.cell(row=row, column=30).value),
                'bb': safe_int(ws.cell(row=row, column=31).value),
                'so': safe_int(ws.cell(row=row, column=32).value),
            }

        # Check if we've gone past the data
        if not game_entry and not pitcher_name and not batter_name:
            # Check next few rows in case of gaps
            has_more = False
            for check_row in range(row + 1, row + 5):
                if (ws.cell(check_row, 1).value or
                    ws.cell(check_row, 8).value or
                    ws.cell(check_row, 26).value):
                    has_more = True
                    break
            if not has_more:
                break

        row += 1

    if verbose:
        print(f"  Found {len(games)} games")
        print(f"  Found {len(batting_players)} batters")
        print(f"  Found {len(pitching_players)} pitchers")

    # Create players from the roster (we still want to know who's on the team)
    player_ids = {}
    for name in set(list(batting_players.keys()) + list(pitching_players.keys())):
        jersey = None
        if name in batting_players:
            jersey = batting_players[name].get('jersey')
        elif name in pitching_players:
            jersey = pitching_players[name].get('jersey')

        player_ids[name] = get_or_create_player(name, format_jersey(jersey))

    # Import games with just the game info (W/L, score, opponent)
    # The Excel only has season totals for batting/pitching, not per-game stats
    # Per-game batting stats should come from CSV imports
    if games:
        for game_date, opponent, w_l, rf, ra in games:
            game_id = create_game(
                season_id=season_id,
                game_date=game_date,
                opponent_name=opponent,
                win_loss=w_l,
                runs_for=rf,
                runs_against=ra
            )
            results['games_imported'] += 1

    # NOTE: We no longer create "Season Totals" placeholder games.
    # The Excel only has aggregated season stats, not per-game breakdowns.
    # Per-game batting/pitching stats should be imported from CSV files.
    # Season totals are calculated on-the-fly from per-game data.

    if verbose:
        print(f"  Note: Season batting/pitching totals not imported (Excel has aggregates only)")
        print(f"  Import per-game stats from CSV files for accurate player stats")

    return results


def import_multi_season_sheet(ws, team_id: int = None, verbose: bool = True) -> List[Dict]:
    """
    Import a sheet with multiple seasons (like 'Fall & Spring')
    These have multiple season blocks stacked vertically
    """
    all_results = []

    # Find all season headers in the sheet
    # Filter out duplicate season entries (same season appearing on consecutive rows)
    raw_season_rows = []
    for row in range(1, 100):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val:
            season_type, year = parse_season_from_text(str(cell_val))
            if season_type and year:
                raw_season_rows.append((row, season_type, year))

    # Deduplicate: only keep first occurrence of each unique season
    season_rows = []
    seen_seasons = set()
    for row, season_type, year in raw_season_rows:
        key = (season_type, year)
        if key not in seen_seasons:
            season_rows.append((row, season_type, year))
            seen_seasons.add(key)

    if verbose:
        print(f"Found {len(season_rows)} season blocks")

    # For each season block, determine its extent and import
    for i, (start_row, season_type, year) in enumerate(season_rows):
        # End row is either next season's start or end of data
        if i + 1 < len(season_rows):
            end_row = season_rows[i + 1][0] - 1
        else:
            end_row = 500  # Safety limit

        if verbose:
            print(f"\nProcessing {season_type} {year} (rows {start_row}-{end_row})")

        # Import this season block
        result = import_season_block(ws, start_row, end_row, season_type, year, team_id, verbose)
        all_results.append(result)

    return all_results


def import_season_block(ws, start_row: int, end_row: int,
                        season_type: str, year: int,
                        team_id: int = None, verbose: bool = True) -> Dict:
    """Import a single season block from a multi-season sheet"""
    results = {
        'season': f"{season_type} {year}",
        'season_id': None,
        'games_imported': 0,
        'batting_records': 0,
        'pitching_records': 0,
        'errors': []
    }

    season_id = get_or_create_season(year, season_type, team_id)
    results['season_id'] = season_id

    # Find the data start row (usually 2-3 rows after season header)
    data_start = start_row + 3

    batting_players = {}
    pitching_players = {}
    games = []

    # Different column layout for 'Fall & Spring' sheet
    # Pitching: G-U (7-21), player in H (8)
    # Batting: W-AE (23-31), player in X (24)

    for row in range(data_start, end_row + 1):
        # Read game info
        game_entry = ws.cell(row=row, column=1).value
        game_date, opponent = parse_game_entry(game_entry)
        if game_date and opponent:
            w_l = ws.cell(row=row, column=2).value
            rf = safe_int(ws.cell(row=row, column=3).value)
            ra = safe_int(ws.cell(row=row, column=4).value)
            games.append((game_date, opponent, w_l, rf, ra))

        # Read pitching stats
        pitcher_name_raw = ws.cell(row=row, column=8).value
        pitcher_jersey_raw = ws.cell(row=row, column=7).value
        pitcher_name = clean_player_name(pitcher_name_raw)
        if pitcher_name and not is_header_or_invalid_player(pitcher_name_raw, pitcher_jersey_raw):
            pitching_players[pitcher_name] = {
                'jersey': pitcher_jersey_raw,
                'app': safe_int(ws.cell(row=row, column=9).value),
                'ip': safe_float(ws.cell(row=row, column=10).value),
                'h': safe_int(ws.cell(row=row, column=11).value),
                'r': safe_int(ws.cell(row=row, column=12).value),
                'k': safe_int(ws.cell(row=row, column=14).value),
                'bb': safe_int(ws.cell(row=row, column=16).value),
                'hbp': safe_int(ws.cell(row=row, column=18).value),
                'pitches': safe_int(ws.cell(row=row, column=19).value),
                'strikes': safe_int(ws.cell(row=row, column=20).value),
            }

        # Read batting stats (columns W-AE, player in X which is 24)
        batter_name_raw = ws.cell(row=row, column=24).value
        batter_jersey_raw = ws.cell(row=row, column=23).value
        batter_name = clean_player_name(batter_name_raw)
        if batter_name and not is_header_or_invalid_player(batter_name_raw, batter_jersey_raw):
            batting_players[batter_name] = {
                'jersey': batter_jersey_raw,
                'ab': safe_int(ws.cell(row=row, column=25).value),
                'r': safe_int(ws.cell(row=row, column=26).value),
                'h': safe_int(ws.cell(row=row, column=27).value),
                'rbi': safe_int(ws.cell(row=row, column=28).value),
                'bb': safe_int(ws.cell(row=row, column=29).value),
                'so': safe_int(ws.cell(row=row, column=30).value),
            }

    if verbose:
        print(f"  Found {len(games)} games, {len(batting_players)} batters, {len(pitching_players)} pitchers")

    # Create players from the roster
    player_ids = {}
    for name in set(list(batting_players.keys()) + list(pitching_players.keys())):
        jersey = None
        if name in batting_players:
            jersey = batting_players[name].get('jersey')
        elif name in pitching_players:
            jersey = pitching_players[name].get('jersey')
        player_ids[name] = get_or_create_player(name, format_jersey(jersey))

    # Create games with just game info (W/L, score, opponent)
    for game_date, opponent, w_l, rf, ra in games:
        create_game(
            season_id=season_id,
            game_date=game_date,
            opponent_name=opponent,
            win_loss=w_l,
            runs_for=rf,
            runs_against=ra
        )
        results['games_imported'] += 1

    # NOTE: We no longer create "Season Totals" placeholder games.
    # Per-game batting/pitching stats should be imported from CSV files.

    if verbose:
        print(f"  Note: Season batting/pitching totals not imported (Excel has aggregates only)")

    return results


def import_workbook(file_path: str, verbose: bool = True) -> Dict:
    """
    Import an entire workbook.
    Extracts team info from the first sheet and creates age group + team.
    """
    wb = load_workbook(file_path, data_only=True)

    all_results = {
        'file': file_path,
        'sheets_processed': 0,
        'total_games': 0,
        'total_batting': 0,
        'total_pitching': 0,
        'team_info': None,
        'team_id': None,
        'details': []
    }

    # Extract team info from first sheet
    first_ws = wb[wb.sheetnames[0]]
    team_info = extract_team_info(first_ws)
    all_results['team_info'] = team_info

    if verbose and team_info.get('team_name'):
        print(f"Team: {team_info['team_name']}")
        print(f"Age Group: {team_info.get('age_group', 'Unknown')}")
        print(f"Location: {team_info.get('location', 'Unknown')}")

    # Create age group and team
    team_id = get_or_create_team_from_info(team_info, verbose)
    all_results['team_id'] = team_id

    for sheet_name in wb.sheetnames:
        if verbose:
            print(f"\n{'='*50}")
            print(f"Processing sheet: {sheet_name}")
            print('='*50)

        ws = wb[sheet_name]

        # Extract team info for this sheet (might be different team per sheet)
        sheet_team_info = extract_team_info(ws)
        sheet_team_id = team_id  # Default to workbook team

        # If this sheet has different team info, create that team too
        if sheet_team_info.get('team_name') and sheet_team_info['team_name'] != team_info.get('team_name'):
            sheet_team_id = get_or_create_team_from_info(sheet_team_info, verbose)

        # Determine sheet type based on content
        # Check if it has multiple UNIQUE seasons (e.g., both Spring and Fall)
        seasons_found = set()
        for row in range(1, 100):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val:
                season_type, year = parse_season_from_text(str(cell_val))
                if season_type and year:
                    seasons_found.add((season_type, year))

        if len(seasons_found) > 1:
            # Multi-season sheet (has different seasons like Spring 2025 AND Fall 2024)
            results = import_multi_season_sheet(ws, sheet_team_id, verbose)
            for r in results:
                all_results['details'].append(r)
                all_results['total_games'] += r['games_imported']
                all_results['total_batting'] += r['batting_records']
                all_results['total_pitching'] += r['pitching_records']
        else:
            # Single season sheet (may have same season mentioned twice)
            result = import_fall_sheet(ws, sheet_team_id, verbose)
            all_results['details'].append(result)
            all_results['total_games'] += result['games_imported']
            all_results['total_batting'] += result['batting_records']
            all_results['total_pitching'] += result['pitching_records']

        all_results['sheets_processed'] += 1

    wb.close()
    return all_results


if __name__ == "__main__":
    # Test import
    import sys

    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = "test-data/12U Teams.xlsx"

    print(f"Importing: {file_path}")
    results = import_workbook(file_path)

    print("\n" + "="*50)
    print("IMPORT SUMMARY")
    print("="*50)
    print(f"Sheets processed: {results['sheets_processed']}")
    print(f"Total games: {results['total_games']}")
    print(f"Total batting records: {results['total_batting']}")
    print(f"Total pitching records: {results['total_pitching']}")
