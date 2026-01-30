"""
Unified file importer for softball stats
Handles both Excel workbooks and GameChanger CSV files
"""

import csv
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook

from database import (
    get_or_create_season, get_or_create_player, create_game,
    add_batting_stats, add_pitching_stats, get_db,
    get_or_create_age_group, get_or_create_our_team
)


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def safe_int(val) -> int:
    """Safely convert value to int"""
    if val is None or val == '':
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def safe_float(val) -> float:
    """Safely convert value to float"""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def format_jersey(val) -> Optional[str]:
    """Format jersey number as whole number string, or None if invalid"""
    if val is None or val == '':
        return None
    str_val = str(val).strip()
    if str_val.upper() in ['#', 'NO', 'NO.', 'NUM', 'NUMBER', 'JERSEY']:
        return None
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return str_val if str_val else None


def clean_player_name(name: str) -> str:
    """Remove handedness indicators and clean player name. Returns empty for headers."""
    if not name:
        return ""
    name = str(name).strip()
    if name.upper() in ['PLAYER', 'NAME', 'PLAYERS', '#', '']:
        return ""
    # Remove (R), (L), (S) suffix
    name = re.sub(r'\s*\([RLS]\)\s*$', '', name)
    return name.strip()


def is_header_or_invalid_player(name: str, jersey) -> bool:
    """Check if a row appears to be a header row rather than actual player data"""
    if not name:
        return True
    name_upper = str(name).strip().upper()
    if name_upper in ['PLAYER', 'NAME', 'PLAYERS', 'TOTALS', '#', '']:
        return True
    if jersey is not None:
        jersey_str = str(jersey).strip().upper()
        if jersey_str in ['#', 'NO', 'NO.', 'NUM', 'NUMBER', 'JERSEY']:
            return True
    return False


# =============================================================================
# FILE TYPE DETECTION
# =============================================================================

def detect_file_type(file_path: str) -> str:
    """
    Detect the type of file based on extension and content.
    Returns: 'excel', 'gamechanger_csv', or 'unknown'
    """
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext in ['.xlsx', '.xls']:
        return 'excel'
    elif ext == '.csv':
        # Check CSV content to determine type
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header:
                    # GameChanger CSV has specific column names
                    if 'BoxScoreComponents__playerName' in header:
                        return 'gamechanger_csv'
                    # Check for ag-cell pattern
                    if any('ag-cell' in str(col) for col in header):
                        return 'gamechanger_csv'
        except Exception:
            pass
        return 'csv'  # Generic CSV

    return 'unknown'


# =============================================================================
# CSV IMPORT (supports GameChanger and standard formats)
# =============================================================================

# GameChanger CSV column mapping (obscure column names)
GAMECHANGER_COLUMN_MAP = {
    'BoxScoreComponents__playerName': 'player_name',
    'ag-cell': 'ab',
    'ag-cell 2': 'r',
    'ag-cell 3': 'h',
    'ag-cell 4': 'rbi',
    'ag-cell 5': 'bb',
    'ag-cell 6': 'so',
}

# Standard CSV column mapping (common column name variations)
STANDARD_COLUMN_MAP = {
    # Player stats columns
    'player': 'player_name', 'player_name': 'player_name', 'name': 'player_name', 'playername': 'player_name',
    'ab': 'ab', 'atbats': 'ab', 'at_bats': 'ab',
    'r': 'r', 'runs': 'r',
    'h': 'h', 'hits': 'h',
    'rbi': 'rbi', 'rbis': 'rbi',
    'bb': 'bb', 'walks': 'bb', 'walk': 'bb',
    'so': 'so', 'k': 'so', 'strikeouts': 'so', 'strikeout': 'so',
    'hbp': 'hbp', 'hitbypitch': 'hbp',
    'sac': 'sac', 'sacrifice': 'sac',
    # Game metadata columns
    'date': 'game_date', 'game_date': 'game_date', 'gamedate': 'game_date',
    'time': 'game_time', 'game_time': 'game_time', 'gametime': 'game_time', 'start_time': 'game_time',
    'opponent': 'opponent', 'opp': 'opponent', 'vs': 'opponent', 'opposing_team': 'opponent',
    'result': 'win_loss', 'w/l': 'win_loss', 'wl': 'win_loss', 'win_loss': 'win_loss', 'outcome': 'win_loss',
    'rf': 'runs_for', 'runs_for': 'runs_for', 'runsfor': 'runs_for', 'our_score': 'runs_for',
    'score': 'runs_for', 'us': 'runs_for',
    'ra': 'runs_against', 'runs_against': 'runs_against', 'runsagainst': 'runs_against',
    'opp_score': 'runs_against', 'their_score': 'runs_against', 'them': 'runs_against',
}


def detect_csv_format(headers: List[str]) -> Tuple[str, Dict[str, str]]:
    """Detect whether this is a GameChanger CSV or standard format."""
    # Check for GameChanger format
    if 'BoxScoreComponents__playerName' in headers or 'ag-cell' in headers:
        col_map = {}
        for csv_col, our_col in GAMECHANGER_COLUMN_MAP.items():
            if csv_col in headers:
                col_map[csv_col] = our_col
        return ('gamechanger', col_map)

    # Check for standard format (case-insensitive)
    headers_lower = [h.lower().strip() for h in headers]
    col_map = {}
    for csv_col, our_col in STANDARD_COLUMN_MAP.items():
        for i, h in enumerate(headers_lower):
            if h == csv_col:
                col_map[headers[i]] = our_col
                break

    if col_map:
        return ('standard', col_map)

    return ('unknown', {})


def preview_csv(file_path: str) -> Dict:
    """
    Preview a CSV file to detect what fields are present without importing.
    Returns info about detected columns, sample data, and what's missing.
    """
    result = {
        'file': file_path,
        'format_type': 'unknown',
        'headers': [],
        'row_count': 0,
        'detected_fields': {},
        'missing_fields': [],
        'has_player_stats': False,
        'errors': []
    }

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    except Exception as e:
        result['errors'].append(f"Failed to read CSV: {e}")
        return result

    if not rows:
        result['errors'].append("CSV file is empty")
        return result

    result['headers'] = list(rows[0].keys())
    result['row_count'] = len(rows)

    # Detect format
    format_type, col_map = detect_csv_format(result['headers'])
    result['format_type'] = format_type

    if format_type == 'unknown':
        result['errors'].append(f"Could not detect CSV format")
        return result

    reverse_map = {v: k for k, v in col_map.items()}

    # Check for player stats columns
    result['has_player_stats'] = 'player_name' in reverse_map

    # Extract sample game metadata from first row
    first_row = rows[0]

    # Check each game metadata field
    game_fields = ['game_date', 'game_time', 'opponent', 'win_loss', 'runs_for', 'runs_against']
    for field in game_fields:
        if field in reverse_map:
            value = first_row.get(reverse_map[field], '').strip()
            if value:
                result['detected_fields'][field] = value
        else:
            result['missing_fields'].append(field)

    # Try to extract from filename
    file_info = parse_filename_for_game_info(file_path)
    if file_info.get('game_date') and 'game_date' not in result['detected_fields']:
        result['detected_fields']['game_date_from_filename'] = file_info['game_date']
        if 'game_date' in result['missing_fields']:
            result['missing_fields'].remove('game_date')
    if file_info.get('opponent') and 'opponent' not in result['detected_fields']:
        result['detected_fields']['opponent_from_filename'] = file_info['opponent']
        if 'opponent' in result['missing_fields']:
            result['missing_fields'].remove('opponent')

    return result


def parse_filename_for_game_info(filename: str) -> Dict:
    """
    Extract game info from filename pattern like:
    game1_sep6_vipers.csv -> {game_num: 1, date: '9/6', opponent: 'vipers'}
    """
    info = {
        'game_date': None,
        'opponent': None,
        'game_num': None
    }

    name = Path(filename).stem

    # Try pattern: game{num}_{month}{day}_{opponent}
    match = re.match(r'game(\d+)_([a-z]+)(\d+)_(.+)', name, re.IGNORECASE)
    if match:
        info['game_num'] = int(match.group(1))
        month_str = match.group(2).lower()
        day = match.group(3)
        info['opponent'] = match.group(4).replace('_', ' ').title()

        months = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
            'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
            'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }
        month_num = months.get(month_str[:3], None)
        if month_num:
            info['game_date'] = f"{month_num}/{day}"

    return info


def import_gamechanger_csv(
    file_path: str,
    season_id: int,
    game_date: Optional[str] = None,
    game_time: Optional[str] = None,
    opponent: Optional[str] = None,
    win_loss: Optional[str] = None,
    runs_for: Optional[int] = None,
    runs_against: Optional[int] = None,
    verbose: bool = True
) -> Dict:
    """
    Import a batting stats CSV file (GameChanger or standard format).

    Returns dict with:
        - game_id: ID of created/updated game
        - stats_imported: Number of player stats imported
        - detected_fields: Dict of fields found in CSV
        - missing_fields: List of fields not found (that user may need to provide)
        - errors: List of any errors
    """
    results = {
        'file': file_path,
        'game_id': None,
        'stats_imported': 0,
        'detected_fields': {},
        'missing_fields': [],
        'errors': []
    }

    # Read the CSV first to detect what columns are present
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    except Exception as e:
        results['errors'].append(f"Failed to read CSV: {e}")
        return results

    if not rows:
        results['errors'].append("CSV file is empty")
        return results

    # Detect CSV format and get column mapping
    headers = list(rows[0].keys()) if rows else []
    format_type, col_map = detect_csv_format(headers)

    if format_type == 'unknown' or not col_map:
        results['errors'].append(f"Could not detect CSV format. Headers found: {headers}")
        return results

    # Build reverse mapping: our_col -> csv_col
    reverse_map = {v: k for k, v in col_map.items()}

    if verbose:
        print(f"  Detected format: {format_type}")
        print(f"  Mapped columns: {list(col_map.values())}")

    # Extract game metadata from CSV if columns exist
    # Use first row for game-level data (date, time, opponent, etc.)
    first_row = rows[0] if rows else {}

    # Check for date in CSV
    if 'game_date' in reverse_map:
        csv_date = first_row.get(reverse_map['game_date'], '').strip()
        if csv_date:
            results['detected_fields']['game_date'] = csv_date
            if not game_date:
                game_date = csv_date

    # Check for time in CSV
    if 'game_time' in reverse_map:
        csv_time = first_row.get(reverse_map['game_time'], '').strip()
        if csv_time:
            results['detected_fields']['game_time'] = csv_time
            if not game_time:
                game_time = csv_time

    # Check for opponent in CSV
    if 'opponent' in reverse_map:
        csv_opponent = first_row.get(reverse_map['opponent'], '').strip()
        if csv_opponent:
            results['detected_fields']['opponent'] = csv_opponent
            if not opponent:
                opponent = csv_opponent

    # Check for W/L in CSV
    if 'win_loss' in reverse_map:
        csv_wl = first_row.get(reverse_map['win_loss'], '').strip().upper()
        if csv_wl in ['W', 'L', 'T', 'WIN', 'LOSS', 'TIE']:
            csv_wl = csv_wl[0]  # Normalize to single letter
            results['detected_fields']['win_loss'] = csv_wl
            if not win_loss:
                win_loss = csv_wl

    # Check for runs_for in CSV
    if 'runs_for' in reverse_map:
        csv_rf = first_row.get(reverse_map['runs_for'], '')
        if csv_rf and str(csv_rf).strip().isdigit():
            results['detected_fields']['runs_for'] = int(csv_rf)
            if runs_for is None:
                runs_for = int(csv_rf)

    # Check for runs_against in CSV
    if 'runs_against' in reverse_map:
        csv_ra = first_row.get(reverse_map['runs_against'], '')
        if csv_ra and str(csv_ra).strip().isdigit():
            results['detected_fields']['runs_against'] = int(csv_ra)
            if runs_against is None:
                runs_against = int(csv_ra)

    # Try to extract game info from filename if still not provided
    if not game_date or not opponent:
        file_info = parse_filename_for_game_info(file_path)
        if not game_date and file_info.get('game_date'):
            game_date = file_info['game_date']
            results['detected_fields']['game_date_from_filename'] = game_date
        if not opponent and file_info.get('opponent'):
            opponent = file_info['opponent']
            results['detected_fields']['opponent_from_filename'] = opponent

    # Track missing required/optional fields
    if not game_date:
        results['missing_fields'].append('game_date')
    if not game_time:
        results['missing_fields'].append('game_time')
    if not opponent:
        results['missing_fields'].append('opponent')
    if win_loss is None:
        results['missing_fields'].append('win_loss')
    if runs_for is None:
        results['missing_fields'].append('runs_for')
    if runs_against is None:
        results['missing_fields'].append('runs_against')

    # Date is required - if still missing, return error
    if not game_date:
        results['errors'].append("Game date is required. Please provide a date for this game.")
        return results

    # Opponent defaults to filename if not provided
    if not opponent:
        opponent = Path(file_path).stem

    if verbose:
        print(f"  Game: {game_date} {game_time or ''} vs. {opponent}")
        if results['detected_fields']:
            print(f"  Detected from CSV: {list(results['detected_fields'].keys())}")
        if results['missing_fields']:
            print(f"  Missing (optional): {results['missing_fields']}")

    # Calculate team totals for runs from player stats if not provided
    if runs_for is None:
        runs_col = reverse_map.get('r')
        if runs_col:
            runs_for = sum(safe_int(row.get(runs_col, 0)) for row in rows)

    # Create the game
    game_id = create_game(
        season_id=season_id,
        game_date=game_date,
        game_time=game_time,
        opponent_name=opponent,
        win_loss=win_loss,
        runs_for=runs_for,
        runs_against=runs_against,
        notes=f"Imported from CSV ({format_type}): {Path(file_path).name}"
    )
    results['game_id'] = game_id

    # Import each player's stats using detected column mapping
    player_col = reverse_map.get('player_name')

    for row in rows:
        player_name = row.get(player_col, '').strip() if player_col else ''
        if not player_name:
            continue

        player_id = get_or_create_player(player_name)

        # Extract stats using the reverse mapping
        ab = safe_int(row.get(reverse_map.get('ab', ''), 0))
        r = safe_int(row.get(reverse_map.get('r', ''), 0))
        h = safe_int(row.get(reverse_map.get('h', ''), 0))
        rbi = safe_int(row.get(reverse_map.get('rbi', ''), 0))
        bb = safe_int(row.get(reverse_map.get('bb', ''), 0))
        so = safe_int(row.get(reverse_map.get('so', ''), 0))
        hbp = safe_int(row.get(reverse_map.get('hbp', ''), 0))
        sac = safe_int(row.get(reverse_map.get('sac', ''), 0))

        add_batting_stats(
            game_id=game_id,
            player_id=player_id,
            ab=ab, r=r, h=h, rbi=rbi, bb=bb, so=so, hbp=hbp, sac=sac
        )
        results['stats_imported'] += 1

    if verbose:
        print(f"  Imported {results['stats_imported']} player stats")

    return results


# =============================================================================
# EXCEL IMPORT
# =============================================================================

def parse_season_from_text(text: str) -> Tuple[Optional[str], Optional[int]]:
    """Parse season type and year from text like 'FALL 2025 RECORD: 30-9'"""
    if not text:
        return None, None
    match = re.search(r'(FALL|SPRING)\s+(\d{4})', text.upper())
    if match:
        return match.group(1).capitalize(), int(match.group(2))
    return None, None


def parse_game_entry(text: str) -> Tuple[Optional[str], Optional[str]]:
    """Parse game date and opponent from text like '9/6 vs. NJ Vipers 12U White'"""
    if not text:
        return None, None
    match = re.match(r'^(\d{1,2}/\d{1,2})\s+vs\.?\s+(.+)$', str(text).strip())
    if match:
        return match.group(1), match.group(2).strip()
    return None, None


def extract_team_info(ws) -> Dict:
    """Extract team info from worksheet header rows."""
    team_info = {
        'team_name': None,
        'age_group': None,
        'location': None,
        'full_name': None
    }

    for row in range(1, 10):
        cell_val = ws.cell(row=row, column=1).value
        if not cell_val:
            continue
        cell_str = str(cell_val).strip()

        if 'TEAM NAME:' in cell_str.upper():
            full_name = cell_str.split(':', 1)[1].strip()
            team_info['full_name'] = full_name
            team_info['team_name'] = full_name
            age_match = re.search(r'(\d{1,2}U)', full_name, re.IGNORECASE)
            if age_match:
                team_info['age_group'] = age_match.group(1).upper()

        elif 'LOCATION:' in cell_str.upper():
            team_info['location'] = cell_str.split(':', 1)[1].strip()

    return team_info


def get_or_create_team_from_info(team_info: Dict, verbose: bool = True) -> Optional[int]:
    """Create age group and team from extracted info."""
    if not team_info.get('age_group') or not team_info.get('team_name'):
        return None

    age_group = team_info['age_group']
    sort_order = int(age_group.replace('U', ''))
    age_group_id = get_or_create_age_group(age_group, sort_order)

    if verbose:
        print(f"Age group: {age_group} (ID: {age_group_id})")

    team_id = get_or_create_our_team(
        name=team_info['team_name'],
        age_group_id=age_group_id,
        location=team_info.get('location')
    )

    if verbose:
        print(f"Team: {team_info['team_name']} (ID: {team_id})")

    return team_id


def import_fall_sheet(ws, team_id: int = None, verbose: bool = True) -> Dict:
    """Import the 'Fall' sheet format"""
    results = {
        'season_id': None,
        'games_imported': 0,
        'batting_records': 0,
        'pitching_records': 0,
        'errors': []
    }

    # Get season from row 7
    season_text = ws['A7'].value or ws['B7'].value
    season_type, year = parse_season_from_text(str(season_text) if season_text else "")

    if not season_type or not year:
        for row in range(1, 10):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val:
                season_type, year = parse_season_from_text(str(cell_val))
                if season_type and year:
                    break

    if not season_type or not year:
        results['errors'].append("Could not determine season from sheet")
        return results

    season_id = get_or_create_season(year, season_type, team_id)
    results['season_id'] = season_id

    if verbose:
        print(f"Importing {season_type} {year} (Season ID: {season_id})")

    batting_players = {}
    pitching_players = {}
    games = []

    row = 9
    while row < 500:
        game_entry = ws.cell(row=row, column=1).value

        game_date, opponent = parse_game_entry(game_entry)
        if game_date and opponent:
            w_l = ws.cell(row=row, column=2).value
            rf = safe_int(ws.cell(row=row, column=3).value)
            ra = safe_int(ws.cell(row=row, column=4).value)
            games.append((game_date, opponent, w_l, rf, ra))

        # Read pitching stats
        pitcher_name_raw = ws.cell(row=row, column=8).value
        pitcher_jersey_raw = ws.cell(row=row, column=7).value
        pitcher_name = clean_player_name(pitcher_name_raw)
        if pitcher_name and not is_header_or_invalid_player(pitcher_name_raw, pitcher_jersey_raw):
            pitching_players[pitcher_name] = {
                'jersey': pitcher_jersey_raw,
                'app': safe_int(ws.cell(row=row, column=9).value),
                'ip': safe_float(ws.cell(row=row, column=10).value),
                'h': safe_int(ws.cell(row=row, column=11).value),
                'r': safe_int(ws.cell(row=row, column=13).value),
                'k': safe_int(ws.cell(row=row, column=15).value),
                'bb': safe_int(ws.cell(row=row, column=17).value),
                'hbp': safe_int(ws.cell(row=row, column=20).value),
                'pitches': safe_int(ws.cell(row=row, column=21).value),
                'strikes': safe_int(ws.cell(row=row, column=22).value),
            }

        # Read batting stats
        batter_name_raw = ws.cell(row=row, column=26).value
        batter_jersey_raw = ws.cell(row=row, column=25).value
        batter_name = clean_player_name(batter_name_raw)
        if batter_name and not is_header_or_invalid_player(batter_name_raw, batter_jersey_raw):
            batting_players[batter_name] = {
                'jersey': batter_jersey_raw,
                'ab': safe_int(ws.cell(row=row, column=27).value),
                'r': safe_int(ws.cell(row=row, column=28).value),
                'h': safe_int(ws.cell(row=row, column=29).value),
                'rbi': safe_int(ws.cell(row=row, column=30).value),
                'bb': safe_int(ws.cell(row=row, column=31).value),
                'so': safe_int(ws.cell(row=row, column=32).value),
            }

        if not game_entry and not pitcher_name and not batter_name:
            has_more = False
            for check_row in range(row + 1, row + 5):
                if (ws.cell(check_row, 1).value or
                    ws.cell(check_row, 8).value or
                    ws.cell(check_row, 26).value):
                    has_more = True
                    break
            if not has_more:
                break

        row += 1

    if verbose:
        print(f"  Found {len(games)} games")
        print(f"  Found {len(batting_players)} batters")
        print(f"  Found {len(pitching_players)} pitchers")

    # Create players from the roster
    player_ids = {}
    for name in set(list(batting_players.keys()) + list(pitching_players.keys())):
        jersey = None
        if name in batting_players:
            jersey = batting_players[name].get('jersey')
        elif name in pitching_players:
            jersey = pitching_players[name].get('jersey')
        player_ids[name] = get_or_create_player(name, format_jersey(jersey))

    # Import games with just game info (W/L, score, opponent)
    # The Excel only has season totals for batting/pitching, not per-game stats
    if games:
        for game_date, opponent, w_l, rf, ra in games:
            create_game(
                season_id=season_id,
                game_date=game_date,
                opponent_name=opponent,
                win_loss=w_l,
                runs_for=rf,
                runs_against=ra
            )
            results['games_imported'] += 1

    # NOTE: We no longer create "Season Totals" placeholder games.
    # Per-game batting/pitching stats should be imported from CSV files.

    if verbose:
        print(f"  Note: Season batting/pitching totals not imported (Excel has aggregates only)")

    return results


def import_season_block(ws, start_row: int, end_row: int,
                        season_type: str, year: int,
                        team_id: int = None, verbose: bool = True) -> Dict:
    """Import a single season block from a multi-season sheet"""
    results = {
        'season': f"{season_type} {year}",
        'season_id': None,
        'games_imported': 0,
        'batting_records': 0,
        'pitching_records': 0,
        'errors': []
    }

    season_id = get_or_create_season(year, season_type, team_id)
    results['season_id'] = season_id

    data_start = start_row + 3

    batting_players = {}
    pitching_players = {}
    games = []

    for row in range(data_start, end_row + 1):
        game_entry = ws.cell(row=row, column=1).value
        game_date, opponent = parse_game_entry(game_entry)
        if game_date and opponent:
            w_l = ws.cell(row=row, column=2).value
            rf = safe_int(ws.cell(row=row, column=3).value)
            ra = safe_int(ws.cell(row=row, column=4).value)
            games.append((game_date, opponent, w_l, rf, ra))

        # Read pitching stats
        pitcher_name_raw = ws.cell(row=row, column=8).value
        pitcher_jersey_raw = ws.cell(row=row, column=7).value
        pitcher_name = clean_player_name(pitcher_name_raw)
        if pitcher_name and not is_header_or_invalid_player(pitcher_name_raw, pitcher_jersey_raw):
            pitching_players[pitcher_name] = {
                'jersey': pitcher_jersey_raw,
                'app': safe_int(ws.cell(row=row, column=9).value),
                'ip': safe_float(ws.cell(row=row, column=10).value),
                'h': safe_int(ws.cell(row=row, column=11).value),
                'r': safe_int(ws.cell(row=row, column=12).value),
                'k': safe_int(ws.cell(row=row, column=14).value),
                'bb': safe_int(ws.cell(row=row, column=16).value),
                'hbp': safe_int(ws.cell(row=row, column=18).value),
                'pitches': safe_int(ws.cell(row=row, column=19).value),
                'strikes': safe_int(ws.cell(row=row, column=20).value),
            }

        # Read batting stats
        batter_name_raw = ws.cell(row=row, column=24).value
        batter_jersey_raw = ws.cell(row=row, column=23).value
        batter_name = clean_player_name(batter_name_raw)
        if batter_name and not is_header_or_invalid_player(batter_name_raw, batter_jersey_raw):
            batting_players[batter_name] = {
                'jersey': batter_jersey_raw,
                'ab': safe_int(ws.cell(row=row, column=25).value),
                'r': safe_int(ws.cell(row=row, column=26).value),
                'h': safe_int(ws.cell(row=row, column=27).value),
                'rbi': safe_int(ws.cell(row=row, column=28).value),
                'bb': safe_int(ws.cell(row=row, column=29).value),
                'so': safe_int(ws.cell(row=row, column=30).value),
            }

    if verbose:
        print(f"  Found {len(games)} games, {len(batting_players)} batters, {len(pitching_players)} pitchers")

    # Create players from the roster
    player_ids = {}
    for name in set(list(batting_players.keys()) + list(pitching_players.keys())):
        jersey = None
        if name in batting_players:
            jersey = batting_players[name].get('jersey')
        elif name in pitching_players:
            jersey = pitching_players[name].get('jersey')
        player_ids[name] = get_or_create_player(name, format_jersey(jersey))

    # Create games with just game info (W/L, score, opponent)
    for game_date, opponent, w_l, rf, ra in games:
        create_game(
            season_id=season_id,
            game_date=game_date,
            opponent_name=opponent,
            win_loss=w_l,
            runs_for=rf,
            runs_against=ra
        )
        results['games_imported'] += 1

    # NOTE: We no longer create "Season Totals" placeholder games.
    # Per-game batting/pitching stats should be imported from CSV files.

    if verbose:
        print(f"  Note: Season batting/pitching totals not imported (Excel has aggregates only)")

    return results


def import_multi_season_sheet(ws, team_id: int = None, verbose: bool = True) -> List[Dict]:
    """Import a sheet with multiple seasons"""
    all_results = []

    raw_season_rows = []
    for row in range(1, 100):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val:
            season_type, year = parse_season_from_text(str(cell_val))
            if season_type and year:
                raw_season_rows.append((row, season_type, year))

    # Deduplicate
    season_rows = []
    seen_seasons = set()
    for row, season_type, year in raw_season_rows:
        key = (season_type, year)
        if key not in seen_seasons:
            season_rows.append((row, season_type, year))
            seen_seasons.add(key)

    if verbose:
        print(f"Found {len(season_rows)} season blocks")

    for i, (start_row, season_type, year) in enumerate(season_rows):
        if i + 1 < len(season_rows):
            end_row = season_rows[i + 1][0] - 1
        else:
            end_row = 500

        if verbose:
            print(f"\nProcessing {season_type} {year} (rows {start_row}-{end_row})")

        result = import_season_block(ws, start_row, end_row, season_type, year, team_id, verbose)
        all_results.append(result)

    return all_results


def import_excel_workbook(file_path: str, verbose: bool = True) -> Dict:
    """Import an entire Excel workbook."""
    wb = load_workbook(file_path, data_only=True)

    all_results = {
        'file': file_path,
        'file_type': 'excel',
        'sheets_processed': 0,
        'total_games': 0,
        'total_batting': 0,
        'total_pitching': 0,
        'team_info': None,
        'team_id': None,
        'details': [],
        'errors': []
    }

    # Extract team info from first sheet
    first_ws = wb[wb.sheetnames[0]]
    team_info = extract_team_info(first_ws)
    all_results['team_info'] = team_info

    if verbose and team_info.get('team_name'):
        print(f"Team: {team_info['team_name']}")
        print(f"Age Group: {team_info.get('age_group', 'Unknown')}")
        print(f"Location: {team_info.get('location', 'Unknown')}")

    team_id = get_or_create_team_from_info(team_info, verbose)
    all_results['team_id'] = team_id

    for sheet_name in wb.sheetnames:
        if verbose:
            print(f"\n{'='*50}")
            print(f"Processing sheet: {sheet_name}")
            print('='*50)

        ws = wb[sheet_name]

        sheet_team_info = extract_team_info(ws)
        sheet_team_id = team_id

        if sheet_team_info.get('team_name') and sheet_team_info['team_name'] != team_info.get('team_name'):
            sheet_team_id = get_or_create_team_from_info(sheet_team_info, verbose)

        # Determine sheet type
        seasons_found = set()
        for row in range(1, 100):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val:
                season_type, year = parse_season_from_text(str(cell_val))
                if season_type and year:
                    seasons_found.add((season_type, year))

        if len(seasons_found) > 1:
            results = import_multi_season_sheet(ws, sheet_team_id, verbose)
            for r in results:
                all_results['details'].append(r)
                all_results['total_games'] += r['games_imported']
                all_results['total_batting'] += r['batting_records']
                all_results['total_pitching'] += r['pitching_records']
        else:
            result = import_fall_sheet(ws, sheet_team_id, verbose)
            all_results['details'].append(result)
            all_results['total_games'] += result['games_imported']
            all_results['total_batting'] += result['batting_records']
            all_results['total_pitching'] += result['pitching_records']

        all_results['sheets_processed'] += 1

    wb.close()
    return all_results


# =============================================================================
# UNIFIED IMPORT INTERFACE
# =============================================================================

def import_file(file_path: str, season_id: Optional[int] = None,
                game_date: Optional[str] = None, game_time: Optional[str] = None,
                opponent: Optional[str] = None, win_loss: Optional[str] = None,
                runs_for: Optional[int] = None, runs_against: Optional[int] = None,
                verbose: bool = True) -> Dict:
    """
    Unified import function that detects file type and imports accordingly.

    Args:
        file_path: Path to the file to import
        season_id: Required for CSV imports, optional for Excel (Excel files contain season info)
        game_date: Game date for CSV imports (e.g., "9/6")
        game_time: Game time for CSV imports (e.g., "2:30 PM")
        opponent: Opponent name for CSV imports
        win_loss: Game result ('W', 'L', or 'T')
        runs_for: Our team's score
        runs_against: Opponent's score
        verbose: Print progress info

    Returns:
        Dict with import results including:
        - file_type: 'excel', 'gamechanger_csv', or 'unknown'
        - For Excel: sheets_processed, total_games, total_batting, total_pitching, team_info
        - For CSV: game_id, stats_imported, detected_fields, missing_fields
        - errors: List of any errors encountered
    """
    file_type = detect_file_type(file_path)

    if file_type == 'excel':
        return import_excel_workbook(file_path, verbose)

    elif file_type in ['gamechanger_csv', 'csv']:
        if not season_id:
            return {
                'file': file_path,
                'file_type': file_type,
                'errors': ['Season ID required for CSV imports. Please select a season first.']
            }
        result = import_gamechanger_csv(
            file_path, season_id,
            game_date=game_date, game_time=game_time,
            opponent=opponent, win_loss=win_loss,
            runs_for=runs_for, runs_against=runs_against,
            verbose=verbose
        )
        result['file_type'] = file_type
        return result

    else:
        return {
            'file': file_path,
            'file_type': 'unknown',
            'errors': [f'Unsupported file type: {Path(file_path).suffix}']
        }


def import_multiple_files(file_paths: List[str], season_id: int = None, verbose: bool = True) -> Dict:
    """
    Import multiple files, detecting type for each.

    Args:
        file_paths: List of file paths to import
        season_id: Required for CSV files, optional for Excel
        verbose: Print progress info

    Returns:
        Dict with combined results
    """
    all_results = {
        'files_processed': 0,
        'excel_files': 0,
        'csv_files': 0,
        'total_games': 0,
        'total_batting': 0,
        'total_pitching': 0,
        'total_stats': 0,
        'errors': [],
        'details': []
    }

    for file_path in file_paths:
        result = import_file(file_path, season_id, verbose)
        all_results['details'].append(result)
        all_results['files_processed'] += 1

        if result.get('errors'):
            all_results['errors'].extend(result['errors'])

        file_type = result.get('file_type', 'unknown')

        if file_type == 'excel':
            all_results['excel_files'] += 1
            all_results['total_games'] += result.get('total_games', 0)
            all_results['total_batting'] += result.get('total_batting', 0)
            all_results['total_pitching'] += result.get('total_pitching', 0)

        elif file_type in ['gamechanger_csv', 'csv']:
            all_results['csv_files'] += 1
            if result.get('game_id'):
                all_results['total_games'] += 1
            all_results['total_stats'] += result.get('stats_imported', 0)

    return all_results


if __name__ == "__main__":
    import sys
    from glob import glob

    if len(sys.argv) > 1:
        files = sys.argv[1:]
    else:
        # Find all importable files in test-data
        files = glob("test-data/*.xlsx") + glob("test-data/game*.csv")

    if not files:
        print("No files found to import")
        sys.exit(1)

    print(f"Importing {len(files)} file(s)...")
    print()

    # For CSV files, we need a season_id - use 1 as default for testing
    results = import_multiple_files(files, season_id=1)

    print("\n" + "="*50)
    print("IMPORT SUMMARY")
    print("="*50)
    print(f"Files processed: {results['files_processed']}")
    print(f"  Excel files: {results['excel_files']}")
    print(f"  CSV files: {results['csv_files']}")
    print(f"Total games: {results['total_games']}")
    print(f"Total batting records: {results['total_batting']}")
    print(f"Total pitching records: {results['total_pitching']}")
    print(f"Total CSV stats: {results['total_stats']}")
    if results['errors']:
        print(f"Errors: {len(results['errors'])}")
        for err in results['errors']:
            print(f"  - {err}")